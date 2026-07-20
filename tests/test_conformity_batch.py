"""
Tests for the multi-file conformity matrix batch feature.

Covers:
  1. generate_batch_conformity_excel() — Overview sheet content, per-matrix
     sheets, sheet-name collision handling, chart presence
  2. _safe_sheet_name() truncation/sanitization
  3. Full pipeline on real conformity matrices (skipped if absent)
  4. End-to-end /api/conformity-batch HTTP route (single file, multiple
     files, and a mix of good + unsupported files)
"""
import io
import sys
from pathlib import Path

import pytest

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from app.qa.conformity_report import generate_batch_conformity_excel, _safe_sheet_name


DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "uploads"
GENTEX = DATA_DIR / "Generic_Conformity_Matrix_of_IRDM_TS_Gentex_reponse.ods"
TIANMA = DATA_DIR / "01843_25_00540_PHYS_GEN_DM17F_Conformity_Matrix_v1_FR_EN_TIANMA_20260427.xlsm"
AUMOVIO = DATA_DIR / "01843_25_00540_PHYS_GEN_DM17F_Conformity_Matrix_v1_FR_EN_AUMOVIO_12DM_2026Apr27.xlsm"


def _fake_analysis(file_name, sheet_name="Matrix", total=10, ok=5, nok=3, na=2,
                   items=None, ok_deep_findings=None):
    return {
        "fileName": file_name,
        "sheetName": sheet_name,
        "totalRows": total,
        "stats": {"OK": ok, "NOK": nok, "NA": na, "EMPTY": 0},
        "items": items or [
            {"rowIndex": i, "reqId": f"REQ-{i}", "reference": "", "description": f"Req {i}",
             "conformityCategory": "OK", "comment": "", "version": ""}
            for i in range(total)
        ],
        "okDeepFindings": ok_deep_findings or [],
        "summary": {"total": total, "ok": ok, "nok": nok, "na": na, "empty": 0},
    }


# ── 1. generate_batch_conformity_excel ────────────────────────────

class TestBatchExcelGeneration:

    def test_raises_on_empty_list(self):
        with pytest.raises(ValueError):
            generate_batch_conformity_excel([])

    def test_single_analysis_produces_overview_and_items_sheet(self):
        from openpyxl import load_workbook
        data = generate_batch_conformity_excel([_fake_analysis("spec_a.xlsx")])
        wb = load_workbook(io.BytesIO(data))
        assert "Overview" in wb.sheetnames
        assert any("Items" in n for n in wb.sheetnames)
        ws = wb["Overview"]
        assert ws.cell(row=4, column=2).value == "spec_a.xlsx"
        assert ws.cell(row=4, column=4).value == 10  # Total
        assert ws.cell(row=4, column=5).value == 5   # OK

    def test_multiple_analyses_each_get_a_row_and_sheets(self):
        from openpyxl import load_workbook
        analyses = [
            _fake_analysis("supplier_a.xlsx", total=20, ok=15, nok=5, na=0),
            _fake_analysis("supplier_b.ods", total=8, ok=2, nok=6, na=0),
        ]
        data = generate_batch_conformity_excel(analyses)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Overview"]
        assert ws.cell(row=4, column=2).value == "supplier_a.xlsx"
        assert ws.cell(row=5, column=2).value == "supplier_b.ods"
        assert ws.cell(row=4, column=4).value == 20
        assert ws.cell(row=5, column=4).value == 8
        # 1 Overview + 2 Items sheets (no deep findings)
        assert len(wb.sheetnames) == 3

    def test_deepok_sheet_only_created_when_findings_exist(self):
        from openpyxl import load_workbook
        analyses = [
            _fake_analysis("no_findings.xlsx"),
            _fake_analysis("with_findings.xlsx", ok_deep_findings=[
                {"severity": "error", "reqId": "REQ-1", "reference": "", "conformity": "OK",
                 "signals": ["contradiction"], "comment": "x", "aiComment": "Suspicious"}
            ]),
        ]
        data = generate_batch_conformity_excel(analyses)
        wb = load_workbook(io.BytesIO(data))
        deepok_sheets = [n for n in wb.sheetnames if "DeepOK" in n]
        assert len(deepok_sheets) == 1
        assert "02" in deepok_sheets[0]  # belongs to the 2nd file

    def test_overview_row_count_matches_file_count(self):
        from openpyxl import load_workbook
        analyses = [_fake_analysis(f"f{i}.xlsx") for i in range(5)]
        data = generate_batch_conformity_excel(analyses)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Overview"]
        filled_rows = sum(
            1 for r in range(4, 4 + 5) if ws.cell(row=r, column=2).value
        )
        assert filled_rows == 5

    def test_one_independent_pie_chart_per_file(self):
        """Each uploaded matrix gets its OWN camembert (pie chart) — not
        one chart aggregating every file together."""
        from openpyxl import load_workbook
        from openpyxl.chart import PieChart
        data = generate_batch_conformity_excel([
            _fake_analysis("a.xlsx", total=20, ok=15, nok=5, na=0),
            _fake_analysis("b.xlsx", total=8, ok=2, nok=6, na=0),
            _fake_analysis("c.xlsx", total=10, ok=10, nok=0, na=0),
        ])
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Overview"]
        assert len(ws._charts) == 3
        assert all(isinstance(c, PieChart) for c in ws._charts)

    def test_each_pie_chart_titled_with_its_own_file_name(self):
        from openpyxl import load_workbook
        data = generate_batch_conformity_excel([
            _fake_analysis("supplier_a.xlsx", ok=5, nok=5, na=0),
            _fake_analysis("supplier_b.xlsx", ok=1, nok=9, na=0),
        ])
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Overview"]
        titles = set()
        for chart in ws._charts:
            t = chart.title.tx.rich.p[0].r[0].t if chart.title and chart.title.tx and chart.title.tx.rich else None
            if t:
                titles.add(t)
        assert titles == {"supplier_a.xlsx", "supplier_b.xlsx"}

    def test_each_pie_chart_reflects_only_its_own_files_stats(self):
        """Chart data must NOT be aggregated — file b's chart must show
        only file b's own OK/NOK counts, not the sum with file a."""
        from openpyxl import load_workbook
        data = generate_batch_conformity_excel([
            _fake_analysis("a.xlsx", total=20, ok=15, nok=5, na=0),
            _fake_analysis("b.xlsx", total=8, ok=2, nok=6, na=0),
        ])
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Overview"]
        # Per-file data tables live in columns K/L (11/12), one 6-row block per file
        block_a = {ws.cell(row=r, column=11).value: ws.cell(row=r, column=12).value for r in range(3, 9)}
        block_b = {ws.cell(row=r, column=11).value: ws.cell(row=r, column=12).value for r in range(9, 15)}
        assert block_a.get("OK") == 15 and block_a.get("NOK") == 5
        assert block_b.get("OK") == 2 and block_b.get("NOK") == 6

    def test_sheet_names_valid_length_and_no_truncated_suffix(self):
        """Sheet names must stay <=31 chars AND never cut a suffix word
        mid-word (regression: 'DeepOK' must never become 'DeepO')."""
        from openpyxl import load_workbook
        long_name = "01843_25_00540_PHYS_GEN_DM17F_Conformity_Matrix_v1_FR_EN_VERY_LONG_SUPPLIER_NAME.xlsm"
        analyses = [_fake_analysis(long_name, ok_deep_findings=[
            {"severity": "warning", "reqId": "R1", "reference": "", "conformity": "OK",
             "signals": [], "comment": "", "aiComment": ""}
        ])]
        data = generate_batch_conformity_excel(analyses)
        wb = load_workbook(io.BytesIO(data))
        for name in wb.sheetnames:
            assert len(name) <= 31
        deepok = [n for n in wb.sheetnames if "DeepOK" in n or "DeepO" in n]
        assert any(n.endswith("DeepOK") for n in wb.sheetnames)

    def test_duplicate_file_names_get_unique_sheets(self):
        from openpyxl import load_workbook
        analyses = [_fake_analysis("same_name.xlsx"), _fake_analysis("same_name.xlsx")]
        data = generate_batch_conformity_excel(analyses)
        wb = load_workbook(io.BytesIO(data))
        item_sheets = [n for n in wb.sheetnames if "Items" in n]
        assert len(item_sheets) == 2
        assert len(set(item_sheets)) == 2  # unique names


class TestSafeSheetName:

    def test_truncates_to_max_len(self):
        assert len(_safe_sheet_name("x" * 50, max_len=31)) == 31

    def test_strips_invalid_excel_characters(self):
        name = _safe_sheet_name("bad/name:with*chars[here]?")
        for ch in ["\\", "/", "*", "[", "]", ":", "?"]:
            assert ch not in name

    def test_empty_input_falls_back_to_sheet(self):
        assert _safe_sheet_name("") == "Sheet"


# ── 3. Real conformity matrices ───────────────────────────────────

class TestRealMatrices:

    @pytest.fixture(scope="class")
    def real_analyses(self):
        if not (GENTEX.exists() and TIANMA.exists() and AUMOVIO.exists()):
            pytest.skip("Sample conformity matrices not found in data/uploads/")
        from app.qa.conformity_analyzer import analyze_conformity_matrix, analysis_to_dict
        return [
            analysis_to_dict(analyze_conformity_matrix(str(p), p.name))
            for p in (GENTEX, TIANMA, AUMOVIO)
        ]

    def test_combined_report_contains_all_three_matrices(self, real_analyses):
        from openpyxl import load_workbook
        data = generate_batch_conformity_excel(real_analyses)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Overview"]
        file_names_in_report = {
            ws.cell(row=r, column=2).value for r in range(4, 4 + len(real_analyses))
        }
        assert file_names_in_report == {a["fileName"] for a in real_analyses}

    def test_item_counts_match_original_analyses(self, real_analyses):
        from openpyxl import load_workbook
        data = generate_batch_conformity_excel(real_analyses)
        wb = load_workbook(io.BytesIO(data))
        for a in real_analyses:
            sheet = next(n for n in wb.sheetnames if a["fileName"][:15] in n and "Items" in n) \
                if any(a["fileName"][:15] in n for n in wb.sheetnames) \
                else None
            # Fallback: match by position (files are processed in order)
            idx = real_analyses.index(a) + 1
            sheet = next(n for n in wb.sheetnames if n.startswith(f"{idx:02d} ") and "Items" in n)
            ws = wb[sheet]
            assert ws.max_row - 1 == len(a["items"])  # -1 for header row


# ── 4. HTTP route (/api/conformity-batch) ─────────────────────────

class TestConformityBatchRoute:

    @pytest.fixture(scope="class")
    def client(self):
        pytest.importorskip("fastapi")
        from fastapi.testclient import TestClient
        from app.conformity_server import app
        return TestClient(app)

    def test_single_file_upload(self, client):
        if not GENTEX.exists():
            pytest.skip("Sample matrix not found")
        with open(GENTEX, "rb") as f:
            res = client.post(
                "/api/conformity-batch",
                files=[("files", (GENTEX.name, f, "application/vnd.oasis.opendocument.spreadsheet"))],
            )
        assert res.status_code == 200
        data = res.json()
        assert data["filesAnalyzed"] == 1
        assert data["filesFailed"] == 0
        assert data["reportExcel"]
        assert data["files"][0]["fileName"] == GENTEX.name

    def test_multiple_files_upload_combined_report(self, client):
        if not (GENTEX.exists() and TIANMA.exists()):
            pytest.skip("Sample matrices not found")
        import base64
        from openpyxl import load_workbook

        with open(GENTEX, "rb") as f1, open(TIANMA, "rb") as f2:
            res = client.post(
                "/api/conformity-batch",
                files=[
                    ("files", (GENTEX.name, f1, "application/octet-stream")),
                    ("files", (TIANMA.name, f2, "application/octet-stream")),
                ],
            )
        assert res.status_code == 200
        data = res.json()
        assert data["filesAnalyzed"] == 2
        assert len(data["files"]) == 2

        xlsx_bytes = base64.b64decode(data["reportExcel"])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Overview" in wb.sheetnames
        ws = wb["Overview"]
        names_in_workbook = {ws.cell(row=r, column=2).value for r in (4, 5)}
        assert names_in_workbook == {GENTEX.name, TIANMA.name}

    def test_mixed_good_and_bad_files_partial_success(self, client):
        if not GENTEX.exists():
            pytest.skip("Sample matrix not found")
        with open(GENTEX, "rb") as f1:
            res = client.post(
                "/api/conformity-batch",
                files=[
                    ("files", (GENTEX.name, f1, "application/octet-stream")),
                    ("files", ("not_a_matrix.txt", io.BytesIO(b"hello"), "text/plain")),
                ],
            )
        assert res.status_code == 200
        data = res.json()
        assert data["filesAnalyzed"] == 1
        assert data["filesFailed"] == 1
        assert data["failed"][0]["fileName"] == "not_a_matrix.txt"

    def test_no_files_returns_error(self, client):
        res = client.post("/api/conformity-batch", files=[])
        assert res.status_code in (400, 422)

    def test_all_files_unsupported_returns_422(self, client):
        res = client.post(
            "/api/conformity-batch",
            files=[("files", ("bad.txt", io.BytesIO(b"hello"), "text/plain"))],
        )
        assert res.status_code == 422
