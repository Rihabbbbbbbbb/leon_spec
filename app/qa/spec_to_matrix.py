"""
Spec → Conformity Matrix generator.

Extracts every requirement (ID + description) from a specification document
and fills them into the official CTS conformity matrix template
(data/refs/Conformity_Matrix_Template.xlsx — the "new version" sheet of
Conformity_matrix_history_management_V1_5, macros removed), so suppliers
receive a pre-filled matrix instead of building it by hand.

Template layout (sheet "new version"):
  - rows 1-9: header block (COUNTIF stats, column titles in row 9)
  - column A: requirement description ("Libellé de la dernière version…")
  - column C: requirement identifier ("Numéro de l'exigence")
  - columns D-I: to be filled by the supplier (applicability, commitment,
    comments, PSA status) — left empty on purpose
  - data starts at row 10
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

from app.config import REFS_DIR

TEMPLATE_PATH = REFS_DIR / "Conformity_Matrix_Template.xlsx"

DATA_START_ROW = 10
COL_DESCRIPTION = 1   # A
COL_REQ_ID = 3        # C

# Requirement identifier schemes seen in CTS specs and conformity matrices:
#   REF-PSP-COMP-001 / APP-xxx / GEN-xxx (spec schemes)
#   REQ-0945126 (matrix scheme)
_REQ_ID_RE = re.compile(
    r"\b("
    r"(?:REF|APP|GEN)-[A-Z0-9][A-Z0-9_.-]{2,40}"
    r"|REQ-\d{4,10}"
    r")\b"
)

# CTS specs use both "shall" and "must" for binding requirements
_SHALL_RE = re.compile(r"\b(?:shall|must)\b", re.IGNORECASE)

# Lines that look like headings / boilerplate, not requirements
_NOISE_RE = re.compile(
    r"^(table of|figure \d|page \d|see |cf\.|nota\b|note\s*:)", re.IGNORECASE
)

_MAX_DESC_CHARS = 600
_MIN_DESC_CHARS = 15


@dataclass
class Requirement:
    """One requirement extracted from the spec."""
    req_id: str        # "" when the spec gives no identifier
    text: str          # requirement description / statement
    line_no: int = 0   # source line (traceability)


def _clean_segment(seg: str) -> str:
    """Normalize a text segment extracted from a (possibly table) line."""
    seg = re.sub(r"\s+", " ", seg).strip(" |·-–—\t")
    return seg.strip()


def _best_description(segments: List[str]) -> str:
    """Pick the requirement description among table-cell segments:
    prefer the segment containing 'shall', else the longest one."""
    candidates = [s for s in segments if len(s) >= _MIN_DESC_CHARS]
    if not candidates:
        return ""
    shall_segs = [s for s in candidates if _SHALL_RE.search(s)]
    pool = shall_segs or candidates
    return max(pool, key=len)


# A line that IS a requirement anchor: an ID alone (possibly followed by a
# short applicability code like " C" or "(v)"), e.g. "REQ-0937326  C".
_ANCHOR_RE = re.compile(
    r"^\s*("
    r"(?:REF|APP|GEN)-[A-Z0-9][A-Z0-9_.-]{2,40}"
    r"|REQ-\d{4,10}"
    r")\b[\s.()A-Za-z0-9]{0,12}$"
)

# Table header row repeated before each requirement in CTS specs
_TABLE_HEADER_RE = re.compile(
    r"requirement\s+number.*description\s+of\s+the\s+requirement", re.IGNORECASE
)

# Metadata segments inside requirement tables (safety attributes, PSA refs)
_META_SEGMENT_RE = re.compile(
    r"att_sdf@|psa_comments@|\{\{|\}\}", re.IGNORECASE
)

# Template-instruction examples accidentally left in specs — not real reqs
_TEMPLATE_EXAMPLE_RE = re.compile(
    r"free to modify the example|it is mandatory to write a requirement"
    r"|<do something>|<expected functional performance|shall\s*…\s*$"
    r"|the requirement engineering template shall|<\s*be made of",
    re.IGNORECASE,
)

# Change-history headings ("New requirements:", "Removed requirements: …")
_HISTORY_HEADING_RE = re.compile(
    r"\b(new|removed|modified)\s+requirements?\s*:?\s*$", re.IGNORECASE
)

_BLOCK_MAX_LINES = 30

# A line that is a self-contained INLINE requirement row
# ("APP-ASU-CD-PERF-0001(0) | The ASU must … | [M8]") — such a line always
# ENDS the current anchor block: it belongs to the next requirement.
_INLINE_ROW_RE = re.compile(
    r"^\s*(?:(?:REF|APP|GEN)-[A-Z0-9][A-Z0-9_.-]{2,40}|REQ-\d{4,10})\b[^|]{0,20}\|"
)


def _fix_id_spacing(text: str) -> str:
    """Repair IDs broken by stray spaces anywhere inside them:
    'REF- ASU…', 'REF-ASU-CD- EXINTER -0006(0)' → 'REF-ASU-CD-EXINTER-0006(0)'.
    Requires at least two dash-segments so prose like 'REF - see below'
    is never touched."""
    def _join(m):
        return re.sub(r"\s+", "", m.group(0))
    return re.sub(
        r"\b(?:REF|APP|GEN)(?:\s*-\s*[A-Z0-9_.]{1,20}){2,}",
        _join,
        text,
    )


def _parse_block_description(block_lines: List[str]) -> Tuple[str, str]:
    """
    Extract (internal_ref, description) from a requirement block.

    The block is the flattened table content following an ID anchor:
    header row, internal requirement number (REF-…), safety/PSA metadata,
    then "…}} | <description> | [upstream]" — where the description may
    span several lines between the pipe separators.
    """
    kept = [l for l in block_lines if not _TABLE_HEADER_RE.search(l)]
    joined = "\n".join(kept)

    m = _REQ_ID_RE.search(joined)
    internal_ref = m.group(1) if m else ""

    segments = []
    for seg in joined.split("|"):
        seg_clean = re.sub(r"\s+", " ", seg).strip(" ·-–—\t")
        if not seg_clean or len(seg_clean) < _MIN_DESC_CHARS:
            continue
        if _META_SEGMENT_RE.search(seg_clean):
            continue
        # Upstream-reference cells like "[SSD_AUE]"
        if re.fullmatch(r"\[?[A-Z0-9_ ,;/-]{1,40}\]?", seg_clean):
            continue
        segments.append(seg_clean)

    desc = _best_description(segments)
    return internal_ref, desc


def extract_requirements(text: str) -> List[Requirement]:
    """
    Extract requirements from the full spec text (paragraphs + flattened
    tables, as produced by extract_text_from_file).

    Three mechanisms, in document order:
    1. BLOCK: an anchor line holding just an ID ("REQ-0937326  C") opens a
       requirement block that runs until the next anchor; the description
       (and the internal REF-… number) are parsed from the block's table
       segments — multi-line descriptions between pipes are handled.
    2. INLINE: a line containing both an ID and its text ("REF-X | The
       system shall … | [SSD]") is parsed directly.
    3. SHALL-ONLY: 'shall' statements without any ID are kept with an
       empty identifier so the matrix stays complete.

    Deduplication prefers the occurrence WITH a description: change-history
    mentions ("New requirements: REQ-123") are superseded by the real
    definition found later in the document.
    """
    text = _fix_id_spacing(text)
    lines = text.split("\n")

    requirements: List[Requirement] = []
    by_id: dict = {}
    seen_texts: set = set()
    consumed = [False] * (len(lines) + 1)
    # line index → Requirement that owns that source line (for attaching
    # continuation lines of long multi-line descriptions in pass 2)
    line_owner: dict = {}

    def _add(rid: str, desc: str, line_no: int):
        desc = re.sub(r"\s+", " ", desc).strip()[:_MAX_DESC_CHARS]
        if desc and _TEMPLATE_EXAMPLE_RE.search(desc):
            return
        # A change-history heading is not a description
        if desc and _HISTORY_HEADING_RE.search(desc):
            desc = ""
        if rid:
            if rid in by_id:
                # Prefer the occurrence that has a description
                if desc and not by_id[rid].text:
                    by_id[rid].text = desc
                return
            req = Requirement(req_id=rid, text=desc, line_no=line_no)
            by_id[rid] = req
            requirements.append(req)
        else:
            if not desc:
                return
            key = re.sub(r"\W+", "", desc.lower())[:120]
            if key in seen_texts:
                return
            seen_texts.add(key)
            requirements.append(Requirement(req_id="", text=desc, line_no=line_no))

    # ── Pass 1: anchor blocks ──────────────────────────────────────
    anchor_idx = [
        i for i, l in enumerate(lines) if _ANCHOR_RE.match(l.strip())
    ]

    def _is_history_mention(i: int) -> bool:
        """Anchor listed under a change-history heading ('New requirements:')
        — register the ID but take no description from that region; the
        real definition later in the document will fill it."""
        for k in (i - 1, i - 2):
            if k >= 0 and lines[k].strip():
                return bool(_HISTORY_HEADING_RE.search(lines[k].strip()))
        return False

    # Coalesce DOORS + internal anchors: in CTS requirement tables the
    # "REQ-… C" (DOORS id) anchor is immediately followed by the internal
    # "REF-…" number of the SAME requirement — merge them into one block
    # keyed by the REQ id, with the REF id kept as internal reference.
    merged_into_prev = set()
    for pos in range(len(anchor_idx) - 1):
        i, nxt = anchor_idx[pos], anchor_idx[pos + 1]
        rid = _ANCHOR_RE.match(lines[i].strip()).group(1)
        nid = _ANCHOR_RE.match(lines[nxt].strip()).group(1)
        if rid.startswith("REQ-") and not nid.startswith("REQ-") and nxt - i <= 4:
            merged_into_prev.add(pos + 1)

    for pos, i in enumerate(anchor_idx):
        if pos in merged_into_prev:
            continue
        rid = _ANCHOR_RE.match(lines[i].strip()).group(1)

        if _is_history_mention(i):
            _add(rid, "", i + 1)
            consumed[i] = True
            continue

        # Block ends at the next non-merged anchor
        end = len(lines)
        nxt_pos = pos + 1
        while nxt_pos < len(anchor_idx) and nxt_pos in merged_into_prev:
            nxt_pos += 1
        if nxt_pos < len(anchor_idx):
            end = anchor_idx[nxt_pos]
        end = min(end, i + 1 + _BLOCK_MAX_LINES)

        # …but never swallow a self-contained inline requirement row —
        # those belong to other requirements (they are parsed in pass 2).
        skip_first = 2 if pos + 1 in merged_into_prev else 0
        for j in range(i + 1 + skip_first, end):
            if _INLINE_ROW_RE.match(lines[j]):
                end = j
                break

        block = lines[i + 1:end]
        internal_ref, desc = _parse_block_description(block)
        if internal_ref and internal_ref != rid and desc:
            desc = f"[{internal_ref}] {desc}"
        _add(rid, desc, i + 1)
        owner = by_id.get(rid)
        for j in range(i, end):
            consumed[j] = True
            if owner is not None:
                line_owner[j] = owner

    # ── Pass 2: inline ID lines + shall-only statements ───────────
    # Logic keywords used inside requirement bodies (IF/THEN blocks) —
    # ALLCAPS but NOT section headings.
    _LOGIC_KEYWORDS = {"IF", "THEN", "ELSE", "SI", "ALORS", "AND", "OR",
                       "ET", "OU", "NOT", "ENDIF", "ELSEIF"}

    def _is_boundary(s: str, noise_is_boundary: bool = True) -> bool:
        """A line that starts a new requirement/table/section — it can
        never be the continuation of the previous cell. In continuation
        contexts noise lines ('See picture…', 'Note: …') are cell CONTENT,
        not boundaries."""
        return bool(
            _ANCHOR_RE.match(s)
            or _INLINE_ROW_RE.match(s)
            or _TABLE_HEADER_RE.search(s)
            or (noise_is_boundary and _NOISE_RE.match(s))
            or _HISTORY_HEADING_RE.search(s)
            or (s == s.upper() and len(s) > 3
                and any(c.isalpha() for c in s)
                and s.strip() not in _LOGIC_KEYWORDS)
        )

    n = len(lines)
    i = 0
    # Last captured requirement + the last source line attributed to it —
    # used to attach continuation lines of long multi-line descriptions.
    last_req: Optional[Requirement] = None
    last_req_end = -10
    orphans: List[Requirement] = []

    def _attachable_gap(start: int, stop: int) -> Optional[List[str]]:
        """Return the plain continuation lines between the last requirement
        and a candidate continuation line, or None if any boundary
        (new table row, header, heading, pipes) separates them."""
        if stop - start > 6:
            return None
        gap: List[str] = []
        for k in range(start, stop):
            s = lines[k].strip()
            if not s:
                continue
            if (consumed[k] or _is_boundary(s, noise_is_boundary=False)
                    or "|" in s or _REQ_ID_RE.search(s)):
                return None
            gap.append(s)
        return gap

    while i < n:
        if consumed[i]:
            owner = line_owner.get(i)
            if owner is not None:
                last_req, last_req_end = owner, i
            i += 1
            continue
        line = lines[i].strip()
        if not line or _NOISE_RE.match(line):
            i += 1
            continue

        ids = _REQ_ID_RE.findall(line)
        if ids:
            # Table rows are "requirement number | description | upstream
            # requirement": IDs sitting AFTER the description segment are
            # upstream references to other documents' requirements — they
            # are NOT requirements of this spec and must not become rows.
            raw_segments = line.split("|")
            seg_clean = [_clean_segment(_REQ_ID_RE.sub(" ", s)) for s in raw_segments]
            desc = _best_description(seg_clean)
            if desc and desc in seg_clean:
                desc_seg_idx = seg_clean.index(desc)
            else:
                desc_seg_idx = len(raw_segments)
            pre_ids, post_ids = [], []
            for si, seg in enumerate(raw_segments):
                for rid in _REQ_ID_RE.findall(seg):
                    (pre_ids if si <= desc_seg_idx else post_ids).append(rid)
            # "id | desc | upstream" rows: trailing ids are upstream refs of
            # OTHER requirements → excluded. "desc | id" rows have no leading
            # id: the trailing id IS the requirement's own identifier.
            req_ids = pre_ids if pre_ids else post_ids
            for rid in req_ids:
                _add(rid, desc, i + 1)

            # ── Continuation merging ──
            # When the row has no upstream cell after the description, the
            # description cell is still OPEN: the following lines (numbered
            # methods, second paragraph of the same cell, …) belong to THIS
            # requirement until the next boundary or a "… | [upstream]"
            # closing line. Prevents one requirement from being split into
            # several rows.
            cell_open = desc_seg_idx >= len(raw_segments) - 1
            req_obj = by_id.get(req_ids[0]) if req_ids else None
            if (cell_open and req_obj is not None
                    and req_obj.line_no == i + 1):
                j = i + 1
                appended = 0
                while j < n and appended < 12:
                    nxt = lines[j].strip()
                    if not nxt:
                        j += 1
                        continue  # blank line inside the same cell
                    if consumed[j] or _is_boundary(nxt, noise_is_boundary=False):
                        break
                    if "|" in nxt:
                        pre_part, post_part = nxt.split("|", 1)
                        # "description | GEN-…-041(0)" is a NEW requirement
                        # row (own id in the trailing cell) — leave it for
                        # normal processing, do not absorb it.
                        if (_REQ_ID_RE.search(post_part)
                                and not post_part.lstrip().startswith("[")):
                            break
                        # Closing line of the cell: "…rest of desc | [M20]"
                        pre = _clean_segment(pre_part)
                        if pre and not _REQ_ID_RE.search(pre):
                            req_obj.text = (req_obj.text + " " + pre)[:_MAX_DESC_CHARS]
                            consumed[j] = True
                            j += 1
                        break
                    req_obj.text = (req_obj.text + " " + nxt).strip()[:_MAX_DESC_CHARS]
                    consumed[j] = True
                    appended += 1
                    j += 1
                last_req, last_req_end = req_obj, j - 1
                i = j
                continue
            if req_obj is not None and req_obj.line_no == i + 1:
                last_req, last_req_end = req_obj, i
            i += 1
            continue

        if _SHALL_RE.search(line):
            # Multi-column data rows (parameter tables: "T_Prearm_ASU | 6 |
            # [0;6] | s | description…") are table data, NOT requirements.
            if line.count("|") >= 2:
                i += 1
                continue
            segments = [_clean_segment(s) for s in line.split("|")]
            desc = _best_description(segments)
            if desc:
                # A shall/must statement shortly after the last requirement,
                # separated only by plain continuation lines (IF/THEN,
                # formulas…), is the CONTINUATION of that requirement's long
                # description — merge it instead of creating an id-less row.
                gap = (
                    _attachable_gap(last_req_end + 1, i)
                    if last_req is not None and "|" not in line
                    else None
                )
                if gap is not None:
                    addition = " ".join(gap + [desc])
                    last_req.text = (
                        (last_req.text + " " + addition).strip()[:_MAX_DESC_CHARS]
                    )
                    for k in range(last_req_end + 1, i + 1):
                        consumed[k] = True
                    last_req_end = i
                elif len(desc) >= 30:
                    # Unattachable id-less statement: kept aside — only used
                    # when the whole document defines NO requirement ids
                    # (otherwise it is table prose, not a requirement).
                    orphans.append(Requirement(req_id="", text=desc, line_no=i + 1))
        i += 1

    # Id-less statements become rows ONLY for documents without any
    # requirement identifiers (else the matrix keeps ids exclusively).
    if not by_id:
        for orph in orphans:
            _add("", orph.text, orph.line_no)

    return requirements


def generate_conformity_matrix(
    requirements: List[Requirement],
    spec_name: str = "",
    template_path: Optional[Path] = None,
) -> bytes:
    """
    Fill the conformity matrix template with the extracted requirements.

    Writes only column A (description) and column C (requirement ID) from
    row 10 down — the supplier columns (applicability, commitment, comments,
    PSA status) and the header block with its COUNTIF statistics are left
    exactly as in the template.

    Returns the filled workbook as XLSX bytes (macro-free).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    path = template_path or TEMPLATE_PATH
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Conformity matrix template not found: {path}. "
            "Expected data/refs/Conformity_Matrix_Template.xlsx."
        )

    wb = load_workbook(str(path))
    ws = wb["new version"]

    wrap = Alignment(wrap_text=True, vertical="top")
    row = DATA_START_ROW
    for req in requirements:
        desc_cell = ws.cell(row=row, column=COL_DESCRIPTION, value=req.text)
        desc_cell.alignment = wrap
        id_cell = ws.cell(row=row, column=COL_REQ_ID, value=req.req_id)
        id_cell.alignment = wrap
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def spec_to_matrix(spec_text: str, spec_name: str = "") -> dict:
    """
    Full pipeline: extract requirements from spec text and produce the
    pre-filled conformity matrix.

    Returns a dict with the XLSX bytes and extraction statistics so the
    caller (API/UI) can report exactly what was extracted:
      {
        "xlsxBytes": bytes,
        "requirementsCount": int,
        "withIdCount": int,
        "withoutIdCount": int,
        "sampleIds": [str, ...],
      }
    """
    requirements = extract_requirements(spec_text)
    xlsx_bytes = generate_conformity_matrix(requirements, spec_name)
    with_id = [r for r in requirements if r.req_id]
    return {
        "xlsxBytes": xlsx_bytes,
        "requirementsCount": len(requirements),
        "withIdCount": len(with_id),
        "withoutIdCount": len(requirements) - len(with_id),
        "sampleIds": [r.req_id for r in with_id[:10]],
    }
