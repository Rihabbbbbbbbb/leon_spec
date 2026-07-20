"""
Conformity Matrix Analyzer — intelligent extraction & AI consistency check.

Reads an ODS or XLSX conformity matrix, auto-detects the sheet and the
"Conformité FNR" / "Commentaires FNR" columns (even if names change),
extracts every requirement with its conformity status and comment, then
uses GPT-4o to flag inconsistencies (e.g. status=OK but comment says
"not tested", or status=NOK but comment says "all good").

Output: structured JSON + pie-chart image (base64 PNG) + PDF report.

Designed for the LEON Copilot Studio integration:
  Copilot Studio → Power Automate → Azure Function /api/conformity → this module
"""
from __future__ import annotations

import base64
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ── Spreadsheet reading ────────────────────────────────────────────
def _read_ods(filepath: str) -> List[List[List[str]]]:
    """
    Read an ODS file using odfpy and return a list of sheets.
    Each sheet is a list of rows; each row is a list of cell strings.
    Handles number-columns-repeated, number-rows-repeated, and number-rows-spanned
    (merged cells) attributes correctly.
    """
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = _safe_ods_load(filepath)
    tables = doc.getElementsByType(Table)
    sheets: List[List[List[str]]] = []

    for table in tables:
        rows = table.getElementsByType(TableRow)
        sheet_data: List[List[str]] = []

        # Track rowspan (merged cell) values: {col_index: (value, remaining_rows)}
        # When a cell has numberrowsspanned > 1, its value should be propagated
        # to the same column in subsequent rows.
        rowspan_values: Dict[int, Tuple[str, int]] = {}

        for row in rows:
            # Skip hidden/filtered rows — AutoFilter sets visibility='filter',
            # manually hidden rows have visibility='collapse'.
            # These rows are not visible to the user in the spreadsheet application
            # but are still present in the ODS XML, so odfpy reads them by default.
            visibility = row.getAttribute("visibility") or ""
            if visibility in ("filter", "collapse"):
                # Decrement rowspan counters for skipped rows
                rowspan_values = {ci: (v, r - 1) for ci, (v, r) in rowspan_values.items() if r > 1}
                continue

            cells = row.getElementsByType(TableCell)
            expanded: List[str] = []
            col_idx = 0
            new_rowspan_values: Dict[int, Tuple[str, int]] = {}

            for cell in cells:
                # Extract text from all paragraphs
                ps = cell.getElementsByType(P)
                text_parts: List[str] = []
                for p in ps:
                    for child in p.childNodes:
                        if hasattr(child, "data"):
                            text_parts.append(str(child.data))
                        elif hasattr(child, "firstChild") and child.firstChild and hasattr(child.firstChild, "data"):
                            text_parts.append(str(child.firstChild.data))
                text = " ".join(text_parts).strip()

                repeat = int(cell.getAttribute("numbercolumnsrepeated") or "1")
                # Cap repeat to avoid huge memory usage (empty trailing cells)
                repeat = min(repeat, 500)
                rowspan = int(cell.getAttribute("numberrowsspanned") or "1")

                for _ in range(repeat):
                    # Check if this column has an active rowspan value
                    if col_idx in rowspan_values and rowspan_values[col_idx][1] > 0:
                        # Use the rowspan value if the current cell is empty
                        rs_val, rs_remaining = rowspan_values[col_idx]
                        if not text and rs_val:
                            expanded.append(rs_val)
                        else:
                            expanded.append(text)
                    else:
                        expanded.append(text)

                    # If this cell has rowspan > 1, register it for subsequent rows
                    if rowspan > 1 and text:
                        new_rowspan_values[col_idx] = (text, rowspan - 1)

                    col_idx += 1

            # Merge new rowspan values with existing ones (decrement remaining)
            for ci, (v, r) in rowspan_values.items():
                if ci not in new_rowspan_values and r > 1:
                    new_rowspan_values[ci] = (v, r - 1)
            rowspan_values = new_rowspan_values

            # Handle number-rows-repeated attribute (empty rows can be repeated)
            row_repeat = row.getAttribute("numberrowsrepeated")
            row_repeat = int(row_repeat) if row_repeat else 1
            row_repeat = min(row_repeat, 10000)  # Cap to avoid memory issues
            for _ in range(row_repeat):
                sheet_data.append(expanded)
        sheets.append(sheet_data)

    return sheets


def _read_xlsx(filepath: str) -> List[List[List[str]]]:
    """
    Read an XLSX/Excel file using openpyxl and return a list of sheets.
    Each sheet is a list of rows; each row is a list of cell strings.

    Handles:
    - Hidden rows (AutoFilter, manually hidden) — skipped
    - Hidden columns — skipped (replaced with empty string)
    - Merged cells — values propagated from top-left to all cells in range
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(filepath, data_only=True)
    sheets: List[List[List[str]]] = []

    for ws in wb.worksheets:
        # Detect hidden columns
        hidden_cols: set = set()
        for ci in range(1, ws.max_column + 1):
            col_letter = get_column_letter(ci)
            col_dim = ws.column_dimensions.get(col_letter)
            if col_dim and col_dim.hidden:
                hidden_cols.add(ci)  # 1-indexed

        # Build merged cell value map: (row, col) → value
        # Propagate the top-left cell value to all cells in the merge range
        merged_values: Dict[Tuple[int, int], str] = {}
        for mc in ws.merged_cells.ranges:
            top_left = ws.cell(row=mc.min_row, column=mc.min_col)
            val = str(top_left.value).strip() if top_left.value is not None else ""
            if val:
                for ri in range(mc.min_row, mc.max_row + 1):
                    for ci in range(mc.min_col, mc.max_col + 1):
                        merged_values[(ri, ci)] = val

        sheet_data: List[List[str]] = []
        for row_idx in range(1, ws.max_row + 1):
            # Skip hidden rows (manually hidden or filtered out by AutoFilter)
            row_dim = ws.row_dimensions.get(row_idx)
            if row_dim and row_dim.hidden:
                continue

            row_values: List[str] = []
            for col_idx in range(1, ws.max_column + 1):
                # Skip hidden columns — replace with empty string
                if col_idx in hidden_cols:
                    row_values.append("")
                    continue
                # Check merged cell value first
                if (row_idx, col_idx) in merged_values:
                    row_values.append(merged_values[(row_idx, col_idx)])
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value).strip() if cell.value is not None else ""
                row_values.append(val)
            sheet_data.append(row_values)
        sheets.append(sheet_data)

    return sheets


def _fix_ods_xml(filepath: str) -> str:
    """
    Fix malformed ODS XML (e.g., duplicate attributes) by re-parsing with lxml
    recovery mode.  Returns the path to a temporary fixed ODS file.

    odfpy uses a strict SAX parser that crashes on duplicate attributes.
    lxml's recover=True keeps the last value for duplicates and silently fixes
    other well-formedness issues.
    """
    import zipfile
    import tempfile
    import os
    from lxml import etree

    tmpdir = tempfile.mkdtemp(prefix="leon_ods_fix_")
    tmp_ods = os.path.join(tmpdir, os.path.basename(filepath))

    with zipfile.ZipFile(filepath, "r") as zin:
        with zipfile.ZipFile(tmp_ods, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                # Fix XML files that may have duplicate attributes
                if item.filename.endswith(".xml"):
                    parser = etree.XMLParser(recover=True, huge_tree=True)
                    tree = etree.fromstring(data, parser=parser)
                    if tree is not None:
                        data = etree.tostring(
                            tree, encoding="UTF-8", xml_declaration=True
                        )
                zout.writestr(item, data)

    return tmp_ods


def _safe_ods_load(filepath: str):
    """
    Load an ODS file with odfpy, falling back to a fixed version if the
    original has malformed XML (duplicate attributes, etc.).

    Also suppresses odfpy's internal print() on parse errors, which crashes
    on Windows cp1252 consoles when the XML contains non-cp1252 characters.
    """
    import contextlib
    import io as _io
    from odf.opendocument import load

    # Suppress odfpy's print() on SAX errors (crashes on Windows cp1252)
    devnull = _io.StringIO()
    with contextlib.redirect_stdout(devnull):
        try:
            return load(filepath)
        except Exception:
            pass

    # Fallback: fix the XML and retry
    fixed_path = _fix_ods_xml(filepath)
    with contextlib.redirect_stdout(devnull):
        return load(fixed_path)


def read_spreadsheet(filepath: str) -> Tuple[List[str], List[List[List[str]]]]:
    """
    Read any supported spreadsheet (ODS or XLSX).
    Returns (sheet_names, sheets_data).
    """
    ext = filepath.lower().rsplit(".", 1)[-1]
    if ext == "ods":
        # odfpy doesn't give sheet names directly; extract from tables
        from odf.table import Table
        doc = _safe_ods_load(filepath)
        tables = doc.getElementsByType(Table)
        sheet_names = [t.getAttribute("name") or f"Sheet_{i}" for i, t in enumerate(tables)]
        sheets_data = _read_ods(filepath)
    elif ext in ("xlsx", "xlsm", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        sheets_data = _read_xlsx(filepath)
    else:
        raise ValueError(f"Unsupported file extension: .{ext}")

    return sheet_names, sheets_data


# ── Intelligent column detection ───────────────────────────────────

# Canonical names and their fuzzy variants
_CONFORMITY_PATTERNS = [
    r"conformit[eé]\s*fnr",
    r"conformity\s*fnr",
    r"supplier\s*conformity",
    r"conformit[eé]\s*(supplier|fournisseur)",
    r"statut\s*fnr",
    r"validation\s*fnr",
    r"conformit[eé]\s*(g[eé]n[eé]ral|global)",
    r"conformity\s*matrix",
    r"conformit[eé]\s*(matrix|matrice)",
    r"supplier\s*response",
    r"supplier\s*status",
    r"statut\s*(supplier|fournisseur)",
    r"^ok$",
    r"^nok$",
]

_COMMENT_PATTERNS = [
    r"commentaires?\s*fnr",
    r"comments?\s*fnr",
    r"supplier\s*comments?",
    r"commentaires?\s*(supplier|fournisseur)",
    r"observations?\s*fnr",
    r"supplier\s*remark",
    r"remarks?\s*(supplier|fournisseur)",
    r"supplier\s*note",
    r"commentaires?\s*supplier",
    r"if\s*nok.*commitment",
    r"minimum\s*commitment",
    r"engagement\s*minimum",
]

# Stellantis verdict columns — "Commentaires STELLANTIS" / "Statut STELLANTIS"
# These columns contain the Stellantis-side OK/NOK verdict and must be checked.
_STELLANTIS_VERDICT_PATTERNS = [
    r"commentaires?\s*stellantis",
    r"statut\s*stellantis",
    r"stellantis\s*comments?",
    r"stellantis\s*status",
    r"stellantis\s*remark",
]

# Version applicable column — "Version Version" / "Version appliquée Applied version"
# These match the DOCUMENT version column (e.g., "Version / Version")
_VERSION_PATTERNS = [
    r"^version\s*version$",
    r"^version\s*appliqu",
    r"^applied\s*version$",
    r"^version$",
]

# Version APPLICABLE column — "Version applicable / Applicable version"
# This is the SUPPLIER's applicable version, distinct from the document version.
# Must be detected separately and prioritized over _VERSION_PATTERNS.
_VERSION_APPLICABLE_PATTERNS = [
    r"version\s*applic",            # "Version applicable"
    r"applicable\s*version",        # "Applicable version"
    r"version\s*appliqu",           # "Version appliquée" (also an applied version)
    r"applied\s*version",           # "Applied version"
]

_REQ_ID_PATTERNS = [
    r"req[-_]?\d",
    r"exigence",
    r"requirement",
    r"liste\s*des\s*doc",
    r"r[eé]f[eé]rence",
    r"reference",
]


def _normalize(text: str) -> str:
    """Normalize text for fuzzy matching: lowercase, strip accents, collapse spaces."""
    if not text:
        return ""
    text = text.lower().strip()
    # Remove accents
    replacements = {"é": "e", "è": "e", "ê": "e", "ë": "e",
                    "à": "a", "â": "a", "ä": "a",
                    "ù": "u", "û": "u", "ü": "u",
                    "î": "i", "ï": "i",
                    "ô": "o", "ö": "o",
                    "ç": "c", "ñ": "n"}
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", " ", text)
    return text


def _match_any(text: str, patterns: List[str]) -> bool:
    """Check if normalized text matches any of the patterns."""
    norm = _normalize(text)
    if not norm:
        return False
    for pat in patterns:
        if re.search(pat, norm):
            return True
    return False


def _find_header_row(sheet: List[List[str]], max_scan: int = 50) -> Optional[int]:
    """
    Find the header row by searching for rows containing both
    'Conformité FNR' and 'Commentaires FNR' (or their variants).

    Also detects Gentex-style headers with separate 'OK'/'NOK' columns
    and 'Supplier comment' columns.
    """
    best_row = None
    best_score = 0
    for ri, row in enumerate(sheet[:max_scan]):
        has_conformity = False
        has_comment = False
        has_ok_nok = False
        for cell in row:
            norm = _normalize(cell)
            if norm == "ok" or norm == "nok":
                has_ok_nok = True
            if _match_any(cell, _CONFORMITY_PATTERNS):
                has_conformity = True
            if _match_any(cell, _COMMENT_PATTERNS):
                has_comment = True
        # Gentex-style: OK/NOK columns + Supplier comment
        if has_ok_nok and has_comment:
            return ri
        score = int(has_conformity) + int(has_comment)
        if score > best_score:
            best_score = score
            best_row = ri
        if score == 2:
            return ri  # Perfect match
    return best_row if best_score > 0 else None


def _find_columns(header_row: List[str]) -> Dict[str, List[int]]:
    """
    Find column indices for conformity, comment, Stellantis verdict, and requirement ID columns.
    Returns dict with 'conformity', 'comment', 'stellantis_verdict', 'req_id' keys
    (lists of indices since there may be multiple sets).

    Also detects Gentex-style OK/NOK separate column pairs (where 'OK' and 'NOK'
    are individual column headers, and the conformity is determined by which
    column has a value).

    Version applicable columns are detected separately from document version columns.
    'version_applicable' = "Version applicable / Applicable version" (supplier's applicable version)
    'version' = "Version / Version" (document version)
    """
    conformity_cols: List[int] = []
    comment_cols: List[int] = []
    stellantis_verdict_cols: List[int] = []
    req_id_cols: List[int] = []
    version_cols: List[int] = []
    version_applicable_cols: List[int] = []
    ok_cols: List[int] = []  # Gentex-style: separate OK column
    nok_cols: List[int] = []  # Gentex-style: separate NOK column

    for ci, cell in enumerate(header_row):
        norm = _normalize(cell)
        # Check for exact "OK" and "NOK" column headers (Gentex style)
        if norm == "ok":
            ok_cols.append(ci)
        elif norm == "nok":
            nok_cols.append(ci)
        elif _match_any(cell, _CONFORMITY_PATTERNS):
            conformity_cols.append(ci)
        # Check version APPLICABLE first (before comment) — in some ODS files,
        # the "Version applicable" column header from a sub-header row says
        # "Commentaires FNR", but the actual column contains version data.
        # We detect it via the _VERSION_APPLICABLE_PATTERNS on the real header.
        if _match_any(cell, _VERSION_APPLICABLE_PATTERNS):
            version_applicable_cols.append(ci)
        if _match_any(cell, _COMMENT_PATTERNS):
            comment_cols.append(ci)
        if _match_any(cell, _STELLANTIS_VERDICT_PATTERNS):
            stellantis_verdict_cols.append(ci)
        if _match_any(cell, _REQ_ID_PATTERNS):
            req_id_cols.append(ci)
        if _match_any(cell, _VERSION_PATTERNS):
            version_cols.append(ci)

    # Remove version_applicable_cols from version_cols (they overlap with "Version appliquée")
    version_cols = [v for v in version_cols if v not in version_applicable_cols]

    # Remove version_applicable_cols from comment_cols (in ODS, the sub-header
    # may label the "Version applicable" column as "Commentaires FNR")
    comment_cols = [c for c in comment_cols if c not in version_applicable_cols]

    # If we found OK/NOK separate column pairs, add them to conformity_cols
    # as pairs (ok_col, nok_col) — the extraction logic will handle them
    if ok_cols or nok_cols:
        # Pair them: each OK column with the next NOK column
        for ok_ci in ok_cols:
            conformity_cols.append(ok_ci)
        for nok_ci in nok_cols:
            conformity_cols.append(nok_ci)

    return {
        "conformity": conformity_cols,
        "comment": comment_cols,
        "stellantis_verdict": stellantis_verdict_cols,
        "req_id": req_id_cols,
        "version": version_cols,
        "version_applicable": version_applicable_cols,
        "ok_cols": ok_cols,
        "nok_cols": nok_cols,
    }


def _find_data_start(sheet: List[List[str]], header_row_idx: int,
                     col_mapping: Optional[Dict[str, List[int]]] = None) -> int:
    """
    Find the first data row after the header.
    A data row is one that has a non-empty value in the first column
    (typically a REQ-ID or reference).

    If col_mapping is provided, also checks req_id columns and conformity columns
    (some files like Gentex have col 0 empty but data in col 2+).
    """
    # Determine which columns to check for data presence
    check_cols = [0]  # Always check col 0
    if col_mapping:
        check_cols.extend(col_mapping.get("req_id", []))
        check_cols.extend(col_mapping.get("conformity", []))
        check_cols.extend(col_mapping.get("comment", []))
    # Deduplicate
    check_cols = list(dict.fromkeys(check_cols))

    for ri in range(header_row_idx + 1, len(sheet)):
        row = sheet[ri]
        if not row:
            continue
        # Check if any of the relevant columns has a non-empty value
        found_val = ""
        for ci in check_cols:
            val = row[ci].strip() if ci < len(row) and row[ci] else ""
            if val:
                found_val = val
                break
        if not found_val:
            continue
        if found_val.startswith("Template") or found_val.startswith("STOP"):
            continue
        # Skip summary rows (BLOCK, CONVERGED, etc.)
        if found_val.upper() in ("BLOCK", "CONVERGED", "NOT CONVERGED",
                                 "NO STELLANTIS ANSWER", "NO SUPPLIER ANSWER",
                                 "NB OF REQ. TO CONV."):
            continue
        return ri
    return header_row_idx + 1


# ── Conformity value classification ────────────────────────────────

_OK_VALUES = {"ok", "conforme", "conform", "c", "yes", "oui", "/", "ko→ok"}
_NOK_VALUES = {"nok", "non conforme", "non conform", "nc", "no", "non", "ko"}
_NA_VALUES = {"na", "n/a", "not applicable", "non applicable", "non app",
              "non implémenté", "non implemente", "non implante"}

# Stellantis domain responsibility codes — when these appear alone (without ": ok"),
# they are just domain assignments (EMPTY — no conformity assessment yet).
# When followed by ": ok" they are classified as OK.
_DOMAIN_CODES = {
    "ee", "me", "sw", "od", "ve", "sys", "dq", "tp", "emc", "opt", "cg",
    "fusa", "function safety", "hw", "mechanical", "electrical", "software",
    "touch", "all",
}


def classify_conformity(value: str, is_assessment: bool = True) -> str:
    """
    Classify a conformity value into a normalized category.
    Returns one of: OK, NOK, NA, EMPTY

    Handles Stellantis-specific patterns:
    - "/" = conform (OK)
    - "EE: ok", "SW: ok", "TP: ok", "ME: ok" = domain-specific OK
    - "EE", "SW", "ME", "OD", "VE", "SYS", "DQ" (domain codes without ": ok") = EMPTY (just domain assignment)
    - "NA" = not applicable
    - Single letters A-H = version codes (EMPTY)
    - Uncertain/pending language = NOK (not confirmed conform)

    Args:
        value: The raw conformity value from the spreadsheet cell.
        is_assessment: If True, uncertain/pending language patterns are NOK.
                       If False, they are EMPTY.
    """
    if not value or not value.strip():
        return "EMPTY"

    norm = _normalize(value).strip()

    # Check NOK first (before OK, since "NOK" contains "OK")
    if norm in _NOK_VALUES or norm.startswith("nok"):
        return "NOK"

    # Domain-specific OK patterns: "EE: ok", "SW: ok", "TP: ok", "ME: ok", "OPT: OK"
    # Also "EE: ok SW: ok" (multi-domain), "DQ: ok", "CG 20260316:OK"
    # Also "Glass is okay", "okay"
    if re.search(r"\b(ok|okay)\b", norm):
        # But not if it also contains NOK or negative words
        if "nok" not in norm and "not ok" not in norm and "not conform" not in norm:
            return "OK"

    # Patterns like "DQ: ok,20260413 ME:" or "CG 20260316:OK"
    if re.search(r":\s*ok", norm) and "nok" not in norm:
        return "OK"

    if norm in _OK_VALUES:
        return "OK"
    if norm in _NA_VALUES or norm.startswith("not applicable") or norm.startswith("non applic"):
        return "NA"

    # Stellantis domain codes without ": ok" — just domain assignments (EMPTY)
    # These do NOT indicate non-conformity; they indicate which domain is responsible.
    # The conformity status comes from the primary conformity column (e.g., "/" = OK).
    if norm in _DOMAIN_CODES:
        return "EMPTY"

    # Domain codes with trailing colon (e.g., "DQ:", "EE:") — incomplete assessment (EMPTY)
    norm_stripped = norm.rstrip(":").strip()
    if norm_stripped in _DOMAIN_CODES:
        return "EMPTY"

    # Check for slash-separated domain codes like "ME/VE", "EE/VE/ME" (EMPTY)
    slash_parts = [p.strip() for p in norm.split("/") if p.strip()]
    if len(slash_parts) >= 2 and all(p in _DOMAIN_CODES for p in slash_parts):
        return "EMPTY"

    # Single letter versions (A-H) are version codes, not conformity (EMPTY)
    if len(norm) == 1 and norm in "abcdefgh":
        return "EMPTY"

    return "EMPTY"


# ── Data structures ─────────────────────────────────────────────────

@dataclass
class ConformityItem:
    """A single requirement row from the conformity matrix."""
    row_index: int
    req_id: str = ""
    reference: str = ""
    description: str = ""
    conformity_raw: str = ""
    conformity_category: str = "EMPTY"
    comment: str = ""
    version: str = ""  # Version applicable (from "Version applicable" or "Version" column)
    column_set: int = 0  # which set of conformity/comment columns (0-based)
    needs_review: bool = False  # True for UNKNOWN/STANDBY items needing manual verification
    classification_confidence: str = "high"  # high/medium/low — confidence in the classification


@dataclass
class ConformityAnalysis:
    """Complete analysis result."""
    sheet_name: str = ""
    header_row: int = -1
    data_start_row: int = -1
    total_rows: int = 0
    sheet_total_rows: int = 0  # Total rows in the sheet (for debugging)
    items: List[ConformityItem] = field(default_factory=list)
    # Statistics
    stats: Dict[str, int] = field(default_factory=dict)
    # AI inconsistency findings
    inconsistencies: List[Dict] = field(default_factory=list)
    # AI deep analysis of OK responses (FNR says OK but comment is suspicious)
    ok_deep_findings: List[Dict] = field(default_factory=list)
    # How the deep analysis was performed: "ia", "ia+motifs" or "motifs"
    ok_deep_method: str = ""
    # Column mapping
    column_mapping: Dict[str, List[int]] = field(default_factory=dict)
    # Debug info
    _debug_repeated_rows: int = 0
    _debug_total_row_elements: int = 0
    _debug_total_expanded_rows: int = 0
    # Chart (base64 PNG)
    chart_base64: str = ""
    # Report text
    report_text: str = ""
    # File name
    file_name: str = ""


# ── Main extraction function ───────────────────────────────────────

def _detect_assessment_columns(
    sheet: List[List[str]],
    conformity_cols: List[int],
    data_start: int,
) -> set:
    """
    Detect which conformity columns are 'assessment columns' (contain ': ok' patterns).

    A column is an assessment column if a significant proportion (>10%) of its
    non-empty data cells contain ': ok' patterns (e.g., 'EE: ok', 'ME: ok').
    This indicates the column is used for conformity assessment, where domain
    codes without ': ok' mean NOK.

    Columns without enough ': ok' patterns are 'domain assignment columns' where
    domain codes are just assignments, not conformity statuses.

    Returns a set of column indices that are assessment columns.
    """
    assessment_cols = set()
    for ci in conformity_cols:
        total_non_empty = 0
        ok_count = 0
        for ri in range(data_start, len(sheet)):
            row = sheet[ri]
            val = row[ci].strip() if ci < len(row) else ""
            if val:
                total_non_empty += 1
                norm = _normalize(val)
                # Check for ': ok' pattern (e.g., 'ee: ok', 'me: ok', 'dq: ok')
                if re.search(r":\s*ok", norm) and "nok" not in norm:
                    ok_count += 1
                elif re.search(r"\b(ok|okay)\b", norm) and "nok" not in norm and "not ok" not in norm:
                    ok_count += 1
        # A column is an assessment column if >10% of non-empty cells have ': ok' patterns
        if total_non_empty > 0 and ok_count / total_non_empty > 0.10:
            assessment_cols.add(ci)
    return assessment_cols


# Category priority for combining multiple column sets (higher = worse)
_CATEGORY_PRIORITY = {
    "NOK": 6,
    "NA": 1,
    "EMPTY": 0,
    "OK": -1,
}


def extract_conformity_data(filepath: str, file_name: str = "") -> ConformityAnalysis:
    """
    Main entry point: read a spreadsheet and extract conformity data.

    Auto-detects:
    - The correct sheet (searches all sheets for Conformité FNR columns)
    - The header row
    - The conformity and comment columns (fuzzy matching, handles name changes)
    - The data start row

    Returns a ConformityAnalysis with all items, statistics, and inconsistencies.
    """
    analysis = ConformityAnalysis(file_name=file_name or filepath)

    sheet_names, sheets_data = read_spreadsheet(filepath)

    # Search all sheets for the one with Conformité FNR columns
    best_sheet_idx = -1
    best_header_row = -1
    best_col_mapping = None

    for si, sheet in enumerate(sheets_data):
        header_row = _find_header_row(sheet)
        if header_row is None:
            continue
        col_mapping = _find_columns(sheet[header_row])
        # Prefer sheets with BOTH conformity and comment columns
        if col_mapping["conformity"] and col_mapping["comment"]:
            # Found a sheet with both conformity and comment columns
            best_sheet_idx = si
            best_header_row = header_row
            best_col_mapping = col_mapping
            analysis.sheet_name = sheet_names[si] if si < len(sheet_names) else f"Sheet_{si}"
            break
        # Fallback: sheet with only conformity columns (no comment columns)
        if col_mapping["conformity"] and best_sheet_idx == -1:
            best_sheet_idx = si
            best_header_row = header_row
            best_col_mapping = col_mapping
            analysis.sheet_name = sheet_names[si] if si < len(sheet_names) else f"Sheet_{si}"
            # Don't break — keep looking for a sheet with both columns

    if best_sheet_idx == -1 or best_col_mapping is None or not best_col_mapping["conformity"]:
        raise ValueError(
            "Could not find 'Conformité FNR' and 'Commentaires FNR' columns "
            "in any sheet of the spreadsheet."
        )

    sheet = sheets_data[best_sheet_idx]
    analysis.header_row = best_header_row
    analysis.column_mapping = best_col_mapping
    analysis.data_start_row = _find_data_start(sheet, best_header_row, best_col_mapping)
    analysis.sheet_total_rows = len(sheet)

    # Debug: check for number-rows-repeated in the raw ODS
    # This helps diagnose if odfpy is handling repeated rows correctly
    try:
        from odf.opendocument import load as _load
        from odf.table import Table as _Table, TableRow as _TableRow
        _doc = _load(filepath)
        _tables = _doc.getElementsByType(_Table)
        _repeated_count = 0
        _total_table_row_elements = 0
        _total_expanded_rows = 0
        for _t in _tables:
            for _r in _t.getElementsByType(_TableRow):
                _total_table_row_elements += 1
                _rep = _r.getAttribute("numberrowsrepeated")
                _rep_val = int(_rep) if _rep else 1
                _total_expanded_rows += _rep_val
                if _rep_val > 1:
                    _repeated_count += 1
        analysis._debug_repeated_rows = _repeated_count
        analysis._debug_total_row_elements = _total_table_row_elements
        analysis._debug_total_expanded_rows = _total_expanded_rows
    except Exception:
        analysis._debug_repeated_rows = -1
        analysis._debug_total_row_elements = -1
        analysis._debug_total_expanded_rows = -1

    conformity_cols = best_col_mapping["conformity"]
    comment_cols = best_col_mapping["comment"]
    req_id_cols = best_col_mapping["req_id"]
    stellantis_verdict_cols = best_col_mapping.get("stellantis_verdict", [])
    version_cols = best_col_mapping.get("version", [])
    version_applicable_cols = best_col_mapping.get("version_applicable", [])
    ok_cols = best_col_mapping.get("ok_cols", [])
    nok_cols = best_col_mapping.get("nok_cols", [])

    # Heuristic: detect "comment" columns that actually contain version applicable data.
    # In some ODS files, the sub-header row labels the "Version applicable" column as
    # "Commentaires FNR", but the data is actually version letters (G, E, D, v4, etc.).
    # If a comment column has >60% version-like values (single letters A-I or v\d+),
    # reclassify it as a version_applicable column.
    # This runs regardless of whether version_applicable_cols is already detected,
    # because the ODS sub-header may mislabel multiple columns.
    if comment_cols:
        data_start = _find_data_start(sheet, best_header_row, best_col_mapping)
        for cmi in list(comment_cols):
            total_non_empty = 0
            version_like = 0
            for ri in range(data_start, min(data_start + 100, len(sheet))):
                row = sheet[ri]
                val = row[cmi].strip() if cmi < len(row) else ""
                if val:
                    total_non_empty += 1
                    cnorm = _normalize(val).strip()
                    if (len(cnorm) == 1 and cnorm in "abcdefghi") or re.match(r"^v\d+", cnorm):
                        version_like += 1
            if total_non_empty > 5 and version_like / total_non_empty > 0.60:
                # This comment column is actually a version applicable column
                comment_cols.remove(cmi)
                if cmi not in version_applicable_cols:
                    version_applicable_cols.append(cmi)

    # Heuristic: detect "conformity" columns that actually contain domain codes
    # (SYS, SW, VE, EE, etc.) rather than conformity values (/, OK, NOK).
    # In ODS files where the sub-header mislabels columns, a "Conformité FNR" column
    # in the second set may actually contain domain assignments (SYS, SW, VE).
    # If a conformity column has >50% domain-code values and <10% actual conformity
    # values (/, OK, NOK), reclassify it as a comment column.
    if len(conformity_cols) > 1:
        data_start = _find_data_start(sheet, best_header_row, best_col_mapping)
        for ci in list(conformity_cols):
            total_non_empty = 0
            domain_count = 0
            conformity_count = 0
            for ri in range(data_start, min(data_start + 100, len(sheet))):
                row = sheet[ri]
                val = row[ci].strip() if ci < len(row) else ""
                if val:
                    total_non_empty += 1
                    cnorm = _normalize(val).strip()
                    # Check if it's a domain code (SYS, SW, VE, EE, ME, etc.)
                    if cnorm in _DOMAIN_CODES or cnorm in ("sys", "sw", "ve", "ee", "me", "od", "opt", "cg", "tp"):
                        domain_count += 1
                    # Check if it's an actual conformity value
                    elif cnorm in _OK_VALUES or cnorm in _NOK_VALUES or cnorm in _NA_VALUES:
                        conformity_count += 1
            if total_non_empty > 5 and domain_count / total_non_empty > 0.50 and conformity_count / total_non_empty < 0.10:
                # This conformity column is actually a comment/domain column
                conformity_cols.remove(ci)
                if ci not in comment_cols:
                    comment_cols.append(ci)

    # Gentex-style format: separate OK and NOK columns
    # In this format, the conformity is determined by which column has a value:
    # if OK column is non-empty → OK, if NOK column is non-empty → NOK
    is_gentex_style = bool(ok_cols or nok_cols)

    # Determine the number of column sets (usually 2: first version + second version)
    # If no comment columns, use the number of conformity columns
    if comment_cols:
        num_sets = min(len(conformity_cols), len(comment_cols))
    else:
        num_sets = len(conformity_cols)

    # Detect which conformity columns are assessment columns (have ": ok" patterns)
    # This determines whether domain codes should be classified as NOK or EMPTY
    assessment_cols = _detect_assessment_columns(sheet, conformity_cols, analysis.data_start_row)

    # Find the last row with any data to avoid counting trailing empty rows
    last_data_row = analysis.data_start_row
    for ri in range(analysis.data_start_row, len(sheet)):
        row = sheet[ri]
        if row and any(c.strip() for c in row if c):
            last_data_row = ri

    # Extract data rows — ONE item per row (combining all column sets)
    items: List[ConformityItem] = []
    for ri in range(analysis.data_start_row, last_data_row + 1):
        row = sheet[ri]
        if not row:
            # Empty row within data range — count as EMPTY
            items.append(ConformityItem(
                row_index=ri, req_id="", conformity_category="EMPTY",
            ))
            continue
        if all(not c.strip() for c in row if c):
            # Completely empty row within data range — count as EMPTY
            items.append(ConformityItem(
                row_index=ri, req_id="", conformity_category="EMPTY",
            ))
            continue

        # Get requirement ID (first non-empty req_id column, or first column)
        req_id = ""
        if req_id_cols:
            for rci in req_id_cols:
                if rci < len(row) and row[rci].strip():
                    req_id = row[rci].strip()
                    break
        if not req_id and row:
            req_id = row[0].strip() if row[0] else ""

        # Skip non-data rows (section headers, summary rows)
        if req_id and req_id.upper() in (
            "BLOCK", "CONVERGED", "NOT CONVERGED", "NO STELLANTIS ANSWER",
            "NO SUPPLIER ANSWER", "NB OF REQ. TO CONV.", "STANDBY",
            "DEVIATION", "OK", "NOK", "NA", "ATT_RESP@FULL", "ATT_RESP@SW",
            "ATT_RESP@DEV", "ATT_RESP@NONE", "ATT_RESP@INT", "ATT_RESP@PTF",
            "ATT_RESP@DEV_INT",
        ):
            continue
        if req_id and (req_id.startswith("Template") or req_id.startswith("STOP")):
            continue

        # Skip non-requirement rows: section headers (e.g., "3.2 Applicable documents"),
        # document names (e.g., "STLA DIAGNOSTIC REQUIREMENT STANDARD - UDS"),
        # category headers (e.g., "GEN"), and document references (e.g., "02017_...RSP-...").
        # Only create items for rows that have a REQ-ID OR conformity-related data.
        has_req_id = req_id.startswith("REQ-")
        has_conformity_data = any(
            (row[ci].strip() if ci < len(row) else "")
            for ci in conformity_cols + stellantis_verdict_cols + comment_cols
        )
        if not has_req_id and not has_conformity_data:
            continue

        # Get reference (second column if available)
        reference = row[1].strip() if len(row) > 1 else ""
        # Get description — try col 5 (Stellantis format) or col 3 (Gentex format)
        description = ""
        if len(row) > 5 and row[5].strip():
            description = row[5].strip()
        elif is_gentex_style and len(row) > 3 and row[3].strip():
            description = row[3].strip()

        # ── Gentex-style: OK/NOK separate columns ──
        # In this format, OK and NOK are separate columns.
        # If OK column has a value → classify it (could be OK, NA, etc.)
        # If NOK column has a value → classify it (could be NOK, NA, etc.)
        # NOK column takes priority if both present (but NA overrides both)
        if is_gentex_style:
            conf_raw = ""
            comment = ""
            best_category = "EMPTY"

            # Check OK columns first — classify the value properly
            ok_cat = None
            ok_raw = ""
            for ok_ci in ok_cols:
                ok_val = row[ok_ci].strip() if ok_ci < len(row) else ""
                if ok_val:
                    ok_raw = ok_val
                    ok_cat = classify_conformity(ok_val, is_assessment=False)
                    break

            # Check NOK columns — classify the value properly
            nok_cat = None
            nok_raw = ""
            for nok_ci in nok_cols:
                nok_val = row[nok_ci].strip() if nok_ci < len(row) else ""
                if nok_val:
                    nok_raw = nok_val
                    nok_cat = classify_conformity(nok_val, is_assessment=False)
                    break

            # Determine final category:
            # - NA takes highest priority (if either column says NA, it's NA)
            # - NOK takes priority over OK
            # - If both present, use the more severe (NOK > OK)
            if ok_cat == "NA" or nok_cat == "NA":
                best_category = "NA"
                conf_raw = ok_raw if ok_cat == "NA" else nok_raw
            elif nok_cat == "NOK":
                best_category = "NOK"
                conf_raw = nok_raw
            elif ok_cat == "OK":
                best_category = "OK"
                conf_raw = ok_raw
            elif nok_cat:
                best_category = nok_cat
                conf_raw = nok_raw
            elif ok_cat:
                best_category = ok_cat
                conf_raw = ok_raw

            # Get comment from comment columns
            all_comments = []
            for cmi in comment_cols:
                cval = row[cmi].strip() if cmi < len(row) else ""
                if cval:
                    all_comments.append(cval)
            combined_comment = " | ".join(all_comments) if all_comments else ""

            # Get version — prioritize version_applicable columns over version columns
            version = ""
            if version_applicable_cols:
                for vci in version_applicable_cols:
                    if vci < len(row) and row[vci].strip():
                        version = row[vci].strip()
                        break
            if not version and version_cols:
                for vci in version_cols:
                    if vci < len(row) and row[vci].strip():
                        version = row[vci].strip()
                        break

            if conf_raw or combined_comment:
                item = ConformityItem(
                    row_index=ri,
                    req_id=req_id,
                    reference=reference,
                    description=description[:200],
                    conformity_raw=conf_raw,
                    conformity_category=best_category,
                    comment=combined_comment,
                    version=version,
                    column_set=0,
                    needs_review=(best_category == "EMPTY" and bool(combined_comment)),
                    classification_confidence="medium" if best_category == "EMPTY" else "high",
                )
                items.append(item)
            continue

        # Collect conformity values from all column sets
        conf_values = []  # list of (raw_value, comment, column_index, set_idx)
        for set_idx in range(num_sets):
            ci = conformity_cols[set_idx]
            cmi = comment_cols[set_idx] if set_idx < len(comment_cols) else -1

            conf_raw = row[ci].strip() if ci < len(row) else ""
            comment = row[cmi].strip() if cmi >= 0 and cmi < len(row) else ""

            if conf_raw or comment:
                conf_values.append((conf_raw, comment, ci, set_idx))

        # Also collect from unpaired comment columns (ODS reclassified columns)
        paired_comment_indices = set()
        for set_idx in range(num_sets):
            if set_idx < len(comment_cols):
                paired_comment_indices.add(comment_cols[set_idx])
        for cmi in comment_cols:
            if cmi not in paired_comment_indices:
                cval = row[cmi].strip() if cmi < len(row) else ""
                if cval:
                    conf_values.append(("", cval, cmi, len(conf_values)))

        # Create item even if no conformity values (as long as we have a reqId)
        # This ensures all data rows are counted in the total

        # Determine the overall conformity category
        if not conf_values:
            # No conformity or comment values in the paired columns.
            # But we still need to check Stellantis verdict columns —
            # the Environmental Technical Specification row has no supplier
            # conformity data but Stellantis marked it as OK.
            # Fall through to the Stellantis verdict check below instead
            # of creating an EMPTY item immediately.
            pass
        classified = []  # list of (conf_raw, comment, cat, set_idx)
        for conf_raw, comment, ci, set_idx in conf_values:
            is_assessment = ci in assessment_cols
            cat = classify_conformity(conf_raw, is_assessment=is_assessment)
            classified.append((conf_raw, comment, cat, set_idx))

        # Determine overall category:
        # - If there are assessment columns with values, use worst category across assessment columns
        # - If NO assessment columns, use the FIRST non-empty conformity value (primary column)
        # - This ensures domain assignment columns don't override the primary conformity status
        assessment_cats = [(r, c, cat, si) for (r, c, cat, si) in classified
                           if si < len(conformity_cols) and conformity_cols[si] in assessment_cols]
        non_assessment_cats = [(r, c, cat, si) for (r, c, cat, si) in classified
                               if si < len(conformity_cols) and conformity_cols[si] not in assessment_cols and r]

        # ── STELLANTIS verdict columns (highest priority) ──
        # "Commentaires STELLANTIS" / "Statut STELLANTIS" columns contain the
        # Stellantis-side OK/NOK verdict. If any says NOK, the item is NOK.
        # If any says OK (and no NOK), the item is OK.
        stellantis_nok = False
        stellantis_nok_raw = ""
        stellantis_ok = False
        stellantis_ok_raw = ""
        for sci in stellantis_verdict_cols:
            sval = row[sci].strip() if sci < len(row) else ""
            if sval:
                snorm = _normalize(sval)
                if snorm in _NOK_VALUES or snorm.startswith("nok"):
                    stellantis_nok = True
                    stellantis_nok_raw = sval
                    break
                if not stellantis_ok and (snorm in _OK_VALUES or re.search(r"\bok\b", snorm)):
                    if "nok" not in snorm and "not ok" not in snorm:
                        stellantis_ok = True
                        stellantis_ok_raw = sval

        if stellantis_nok:
            # Stellantis verdict NOK overrides everything
            best_conf_raw = stellantis_nok_raw
            best_comment = ""
            best_category = "NOK"
            best_set_idx = 0
        elif non_assessment_cats:
            # Non-assessment (primary) columns take priority — they contain the
            # definitive conformity status (e.g., "/" = OK).
            # Use the FIRST non-empty value (primary column).
            best = non_assessment_cats[0]
            best_conf_raw = best[0]
            best_comment = best[1]
            best_category = best[2]
            best_set_idx = best[3]
            # ── Stellantis OK fallback ──
            # If the non-assessment column only has a domain code (EMPTY), but
            # Stellantis verdict says OK, use the Stellantis OK.
            # This fixes rows like the Environmental Technical Specification
            # where the supplier conformity is a domain code (e.g., 'VE')
            # but Stellantis marked it as OK.
            if best_category == "EMPTY" and stellantis_ok:
                best_conf_raw = stellantis_ok_raw
                best_comment = best_comment  # keep the original comment
                best_category = "OK"
                best_set_idx = 0
        elif assessment_cats:
            # No primary column value — use assessment columns
            best = max(assessment_cats, key=lambda x: _CATEGORY_PRIORITY.get(x[2], 0))
            best_conf_raw = best[0]
            best_comment = best[1]
            best_category = best[2]
            best_set_idx = best[3]
            # ── Stellantis OK fallback ──
            # Same fallback as above for assessment columns
            if best_category == "EMPTY" and stellantis_ok:
                best_conf_raw = stellantis_ok_raw
                best_comment = best_comment  # keep the original comment
                best_category = "OK"
                best_set_idx = 0
        elif stellantis_ok:
            # No supplier conformity value, but Stellantis says OK
            best_conf_raw = stellantis_ok_raw
            best_comment = ""
            best_category = "OK"
            best_set_idx = 0
        else:
            best_conf_raw = ""
            best_comment = ""
            best_category = "EMPTY"
            best_set_idx = 0

        # Extract version applicable — PRIORITIZE "Version applicable" columns
        # over "Version" (document version) columns.
        # The "Version applicable" column (e.g., Col H in Stellantis format)
        # contains the supplier's applicable version, which is what the user wants.
        # The "Version" column (e.g., Col D) contains the document version.
        version = ""
        # 1st priority: version_applicable columns ("Version applicable / Applicable version")
        if version_applicable_cols:
            for vci in version_applicable_cols:
                if vci < len(row) and row[vci].strip():
                    version = row[vci].strip()
                    break
        # 2nd priority: version columns (document version — "Version / Version")
        if not version and version_cols:
            for vci in version_cols:
                if vci < len(row) and row[vci].strip():
                    version = row[vci].strip()
                    break
        # 3rd priority: extract from supplier comment (version is sometimes
        # duplicated in the comment column, especially in ODS files where
        # the "Version applicable" column is mislabeled as "Commentaires FNR")
        if not version and conf_values:
            for conf_raw, comment_val, ci, set_idx in conf_values:
                if comment_val and comment_val.strip():
                    cnorm = _normalize(comment_val).strip()
                    # Single letter version codes (A-I) or version patterns (v1.0, v34.0)
                    if (len(cnorm) == 1 and cnorm in "abcdefghi") or re.match(r"^v\d+", cnorm):
                        version = comment_val.strip()
                        break

        # Combine all comments from all column sets, filtering out version values
        # (the version is often duplicated in the comment column — we want actual comments only)
        all_comments = []
        for (_, c, _, _) in conf_values:
            if c and c.strip():
                # Skip if this comment is just the version value (e.g., 'G', 'v1.0')
                if version and c.strip() == version:
                    continue
                # Also skip if the comment is just a single letter (version code)
                # that matches the version
                cnorm = _normalize(c).strip()
                if version and cnorm == _normalize(version).strip():
                    continue
                all_comments.append(c.strip())
        # Note: unpaired comment columns are already in conf_values (added above),
        # so they are collected in the loop above. No separate collection needed.
        # Deduplicate comments (ODS may have same value in multiple comment columns)
        seen = set()
        unique_comments = []
        for c in all_comments:
            if c not in seen:
                seen.add(c)
                unique_comments.append(c)
        combined_comment = " | ".join(unique_comments) if unique_comments else best_comment
        # Also filter best_comment if it's just the version
        if best_comment and version:
            if best_comment.strip() == version or _normalize(best_comment).strip() == _normalize(version).strip():
                combined_comment = " | ".join(unique_comments) if unique_comments else ""

        # Determine if this item needs manual review and its confidence level
        needs_review = False
        confidence = "high"
        if best_category == "EMPTY" and combined_comment:
            # Empty conformity but has a comment — might need review
            needs_review = True
            confidence = "medium"

        item = ConformityItem(
            row_index=ri,
            req_id=req_id,
            reference=reference,
            description=description[:200],
            conformity_raw=best_conf_raw,
            conformity_category=best_category,
            comment=combined_comment,
            version=version,
            column_set=best_set_idx,
            needs_review=needs_review,
            classification_confidence=confidence,
        )
        items.append(item)

    analysis.items = items
    analysis.total_rows = len(items)

    # Compute statistics
    stats: Dict[str, int] = {}
    for item in items:
        cat = item.conformity_category
        stats[cat] = stats.get(cat, 0) + 1
    analysis.stats = stats

    return analysis


# ── AI inconsistency detection ──────────────────────────────────────

# ── Negative-signal patterns (contradict OK status) ──
# Each pattern is (regex, label, weight) — weight contributes to a severity score.
_NEGATIVE_SIGNALS: List[tuple] = [
    # Direct negation of conformity (allow optional 'be' between not and verb)
    (r"\bnot?\s*(?:be\s+)?(ok|conform|test|implement|done|complete|met|satisf)",
     "negated_conformity", 3),
    (r"\bnok\b", "nok", 3),
    (r"\bko\b", "ko", 3),
    (r"\bfail(ed|ure)?\b", "fail", 3),
    (r"\berror\b", "error", 2),
    (r"\bdefect\b", "defect", 3),
    (r"\bbug\b", "bug", 2),
    (r"\bbroken\b", "broken", 3),
    (r"\bmissing\b", "missing", 2),
    (r"\bincomplete\b", "incomplete", 3),
    (r"\bnon\s*(conform|ok|test|verif|implement)\b", "non_conformity", 3),

    # Cannot / unable / impossible / can not
    (r"\b(cannot|can'?t|can\s+not|unable|impossible|no\s+way|no\s+solution)\b",
     "cannot", 3),
    (r"\b(pas\s+(possible|capable|en\s+mesure))\b", "fr_cannot", 3),

    # Does not meet / comply / satisfy
    (r"\b(does\s+not|do\s+not|doesn'?t|don'?t)\s*(meet|comply|satisf|fulfil)",
     "not_meet", 3),
    (r"\bnot\s+(met|satisfied|achieved|fulfilled|compliant)\b", "not_met", 3),

    # Problem / issue / concern / risk
    (r"\b(problem|issue|concern|risk|trouble|difficulty)\b", "problem", 2),
    (r"\b(problème|souci|préoccupation|risque)\b", "fr_problem", 2),

    # Pending / waiting / blocked / not ready
    (r"\b(pending|wait|block|not\s+ready|not\s+done|not\s+available)\b",
     "pending", 2),
    (r"\b(en\s+cours|en\s+attente|à\s+venir|à\s+faire|à\s+vérifier|à\s+confirmer|à\s+définir)\b",
     "fr_pending", 2),

    # Deviation / derogation / waiver / exception
    (r"\b(deviation|derogation|waiver|exception|exemption|dispense)\b",
     "deviation", 2),
    (r"\b(dérogation|écart|dispense)\b", "fr_deviation", 2),

    # Partial / limited / workaround
    (r"\b(partial|partially|limitation|limited|workaround|interim|temporary)\b",
     "partial", 2),
    (r"\b(partiel|partielle|limité|limitée|solution\s+de\s+rechange)\b",
     "fr_partial", 2),

    # TODO / TBD / TBA / to be defined
    (r"\b(todo|tbd|tba|to\s+be\s+(defined|confirmed|determined|verified|checked))\b",
     "todo", 2),

    # Conflict / contradiction / mismatch / gap
    (r"\b(conflict|contradict|mismatch|discrepancy|gap|inconsisten)\b",
     "conflict", 2),
    (r"\b(conflit|contradiction|écart|incohérent|incohérence)\b",
     "fr_conflict", 2),

    # Rejected / reject
    (r"\b(reject(ed|ion)?|refus(ed|al)?)\b", "rejected", 3),
    (r"\b(refus(é|ée|er))\b", "fr_rejected", 3),

    # Still / remaining / outstanding
    (r"\b(still|remaining|outstanding|not\s+yet)\b", "remaining", 1),
    (r"\b(reste|encore|pas\s+encore)\b", "fr_remaining", 1),

    # Under review / investigation
    (r"\b(under\s+(review|investigation)|being\s+(reviewed|investigated))\b",
     "under_review", 2),

    # French: pas conforme / pas ok / pas terminé
    (r"\bpas\s+(conforme|ok|termin|fait|prêt|pret|valid)\b", "fr_pas", 3),
    (r"\bnon\s*conforme\b", "fr_non_conforme", 3),

    # Instead of / replaced by (supplier proposes alternative — may not meet original req)
    (r"\b(instead\s+of|replaced\s+by|substitut)\b", "instead_of", 1),

    # No X (negation of key nouns)
    (r"\bno\s+(cybersecurity|security|safety|solution|way|support|capability)\b",
     "no_noun", 2),

    # N/A or not applicable in comment when status is OK (not NA)
    (r"\b(n/?a|not\s+applicable|non\s+applicable|hors\s+périmètre|hors\s+scope)\b",
     "na_in_ok", 2),

    # Delay / late / retard
    (r"\b(delay|late|overdue|retard)\b", "delay", 1),
]

# ── Positive-signal patterns (contradict NOK status) ──
# Each pattern is (regex, label, weight).
_POSITIVE_SIGNALS: List[tuple] = [
    (r"\bok\b", "ok", 2),
    (r"\bconform(e|ed)?\b", "conform", 2),
    (r"\bgood\b", "good", 1),
    (r"\bpass(ed|ing)?\b", "pass", 2),
    (r"\bdone\b", "done", 1),
    (r"\bcompleted\b", "complete", 1),
    (r"\bfinish(ed|ing)?\b", "finish", 1),
    (r"\bready\b", "ready", 1),
    (r"\bvalidated\b", "valid", 2),
    (r"\bconforme\b", "fr_conforme", 2),
    (r"\btermin(é|ée|er|e)\b", "fr_termin", 1),
    (r"\bfait\b", "fr_fait", 1),
    (r"\bprêt|pret\b", "fr_pret", 1),
    (r"\bvalide\b", "fr_valide", 2),
    (r"\bsatisf(ied|ies|action)\b", "satisfied", 2),
    (r"\bmeets\b", "meets", 1),
    (r"\bmeeting\s+(the|this|all|requirement|target|spec)", "meets", 1),
    (r"\bcompl(ies|ied|iant)\b", "compliant", 2),
    (r"\bno\s+(issue|problem|defect|error)\b", "no_issue", 2),
]

# ── Domain codes that are NOT real comments (should not trigger inconsistency) ──
_DOMAIN_CODE_RE = re.compile(
    r"^(sys|sw|ve|ee|me|od|opt|cg|tp|hw|mech|dq|fusa|ipm|all)"
    r"(\s*/\s*(sys|sw|ve|ee|me|od|opt|cg|tp|hw|mech|dq|fusa|ipm|all))*\s*$",
    re.IGNORECASE,
)

# ── Patterns that neutralise a positive signal (context matters) ──
# e.g., "not ok" should NOT count as positive "ok"
_POSITIVE_NEUTRALISER = re.compile(
    r"\b(not?|non|pas|no)\s+"
    r"(ok|conform|conforme|good|pass|done|complete|finish|ready|valid|"
    r"termin|fait|prêt|pret|valide|satisf|meet|compl)",
    re.IGNORECASE,
)


def _score_comment(comment: str, signals: List[tuple]) -> List[tuple]:
    """
    Score a comment against a list of signal patterns.
    Returns a list of (label, weight, matched_text) for all matches.
    """
    matches: List[tuple] = []
    cnorm = _normalize(comment)
    for pattern, label, weight in signals:
        m = re.search(pattern, cnorm)
        if m:
            matches.append((label, weight, m.group()))
    return matches


def detect_inconsistencies(analysis: ConformityAnalysis) -> List[Dict]:
    """
    Detect logical non-coherence in supplier OK responses.

    FOCUS: Only analyze items where the supplier marked 'OK' — check if the
    comment contradicts the OK status (e.g. comment describes non-conformity
    while status is OK). This is the core non-coherence pattern that matters:
    status=OK but comment tells a different story.

    Checks performed (OK-items only):
    - status=OK but comment contains negative/non-conform language
    - status=OK but comment mentions N/A or not applicable
    - status=OK but no comment provided

    LLM check (optional): deeper semantic analysis of comment vs status.
    """
    inconsistencies: List[Dict] = []

    for item in analysis.items:
        cat = item.conformity_category
        comment = item.comment.strip()
        conf_raw = item.conformity_raw.strip()

        # Only analyze OK items — the focus is on supplier-declared OK
        # with comments that may reveal hidden non-conformity.
        if cat != "OK":
            continue

        # Skip pure domain-code comments (SYS, SW, VE, EE, etc.)
        is_domain_only = bool(_DOMAIN_CODE_RE.match(comment)) if comment else False

        issue = None

        # ── Check A: OK status but negative/non-conform comment ──
        # This is the CORE non-coherence detection: the supplier says OK
        # but the comment language suggests otherwise.
        if comment and not is_domain_only:
            neg_matches = _score_comment(comment, _NEGATIVE_SIGNALS)
            if neg_matches:
                score = sum(w for _, w, _ in neg_matches)
                labels = ", ".join(sorted(set(l for l, _, _ in neg_matches)))
                matched_texts = [t for _, _, t in neg_matches]
                severity = "error" if score >= 4 else "warning"
                issue = {
                    "type": "OK_NEGATIVE_COMMENT",
                    "severity": severity,
                    "req_id": item.req_id,
                    "conformity": conf_raw,
                    "comment": comment,
                    "score": score,
                    "signals": labels,
                    "matched": matched_texts,
                    "explanation": (
                        f"Le fournisseur a marqué '{conf_raw}' (OK) mais le "
                        f"commentaire contient un langage négatif ou de non-conformité "
                        f"(signaux: {labels}, score: {score}): "
                        f"'{comment[:200]}'. Incohérence logique — le commentaire "
                        f"ne correspond pas au statut OK déclaré."
                    ),
                }

        # ── Check B: OK status but comment mentions N/A or not applicable ──
        # Logical gap: if it's not applicable, why is it marked OK?
        if comment and not is_domain_only and not issue:
            na_match = re.search(
                r"\b(n/?a|not\s+applicable|non\s+applicable|hors\s+périmètre|hors\s+scope)\b",
                _normalize(comment),
            )
            if na_match:
                issue = {
                    "type": "OK_NA_COMMENT",
                    "severity": "warning",
                    "req_id": item.req_id,
                    "conformity": conf_raw,
                    "comment": comment,
                    "signals": "na_in_ok",
                    "matched": [na_match.group()],
                    "explanation": (
                        f"Le fournisseur a marqué '{conf_raw}' (OK) mais le "
                        f"commentaire mentionne 'N/A' ou 'not applicable': "
                        f"'{comment[:200]}'. Incohérence — si l'exigence n'est "
                        f"pas applicable, le statut OK n'est pas cohérent."
                    ),
                }

        # ── Check C: OK status but no comment ──
        # Warning only — OK without explanation is not a logical contradiction
        # but reduces auditability.
        if not comment and item.column_set == 0:
            if not re.search(r"\bok\b", _normalize(conf_raw)):
                issue = {
                    "type": "OK_NO_COMMENT",
                    "severity": "warning",
                    "req_id": item.req_id,
                    "conformity": conf_raw,
                    "comment": "",
                    "explanation": (
                        f"Le fournisseur a marqué '{conf_raw}' (OK) sans "
                        f"commentaire. Un commentaire justifiant la conformité "
                        f"est recommandé pour l'auditabilité."
                    ),
                }

        if issue:
            inconsistencies.append(issue)

    analysis.inconsistencies = inconsistencies
    return inconsistencies


# ── Deep OK analysis (FNR says OK, but comment looks suspicious) ───

# Patterns that, when found in an OK item's comment, suggest hidden non-conformity.
# These go beyond simple negative keywords — they detect ambiguous or worrying
# language that warrants human review.
_OK_SUSPICION_PATTERNS: List[tuple] = [
    # Pending / to be confirmed
    (r"\bto\s+be\s+(confirmed|defined|determined|verified|checked|decided|tested)\b",
     "pending_confirmation", 3),
    (r"\b(in\s+development|in\s+progress|en\s+cours|à\s+confirmer|à\s+définir|à\s+vérifier)\b",
     "in_development", 2),

    # Need / require further action
    (r"\b(need(s|ed)?\s+(to|further|more|additional|clarification|review|check|confirm|investigation|discuss)|"
     r"require(s|d)?\s+(further|more|clarification|review|confirmation)|"
     r"gentex\s+to\s+(check|confirm|provide|review|explain|verify))",
     "needs_action", 2),
    (r"\bstla\s+to\s+(check|confirm|provide|review|quantify|decide|define)",
     "stla_action", 2),
    (r"\b(please\s+(provide|clarify|confirm|check)|"
     r"merci\s+de\s+(confirmer|vérifier|préciser|clarifier))",
     "please_clarify", 2),

    # Not applicable / out of scope but marked OK
    (r"\b(not\s+applicable|non\s+applicable|n/?a|hors\s+scope|hors\s+périmètre|"
     r"not\s+in\s+scope|no\s+cybersecurity|no\s+solution)",
     "na_language", 3),

    # Instead of / replaced by / deviation
    (r"\b(instead\s+of|replaced\s+by|substitut|deviation|derogation|"
     r"dérogation|waiver|workaround|alternate|alternative)\b",
     "alternative_approach", 2),

    # Not responsible / not our scope
    (r"\b(not\s+responsible|not\s+our\s+scope|not\s+in\s+scope|"
     r"pas\s+responsable|pas\s+de\s+notre\s+ressort)\b",
     "not_responsible", 3),

    # Conflict / contradiction
    (r"\b(conflict|contradiction|incompatible|inconsistent|"
     r"conflit|incohérent)\b",
     "conflict", 2),

    # Partial / limited / under review
    (r"\b(partial|partially|limited\s+to|only\s+for|except\s+for|"
     r"under\s+review|under\s+investigation|"
     r"partiel|partielle|limité\s+à)\b",
     "partial_limited", 2),

    # Cannot / unable
    (r"\b(cannot|can'?t|can\s+not|unable\s+to|not\s+possible|impossible|"
     r"pas\s+possible|impossible\s+de)\b",
     "cannot", 3),

    # Follow same with / same as (may be OK but needs verification)
    (r"\b(follow\s+(same|the\s+same)\s+(as|with)|same\s+as\s+previous)",
     "follow_same", 1),

    # Exception / unless
    (r"\b(exception|unless|sauf|sous\s+réserve|under\s+condition|provided\s+that)\b",
     "conditional", 1),

    # Temporarily / interim / for now
    (r"\b(temporar|interim|for\s+now|provisional|provisoire|temporaire)\b",
     "temporary", 2),

    # Target is X (suggests target not yet met)
    (r"\btarget\s+is\b", "target_is", 1),

    # Risk / concern
    (r"\b(risk|concern|attention|caution|warning|"
     r"risque|préoccupation|attention\s+à)\b",
     "risk_concern", 1),

    # Remaining / outstanding / still
    (r"\b(still\s+(to|need|pending|missing|remaining|outstanding)|"
     r"reste\s+à|encore\s+à)\b",
     "remaining", 2),

    # Discuss / discussion / meeting
    (r"\b(discuss(ed|ion)?\s+(in|with|needed|required)|"
     r"meeting\s+to\s+be\s+organized|"
     r"discuté|à\s+discuter)\b",
     "discussion_needed", 1),

    # Should / shall / must (normative but in comment means not yet done)
    (r"\b(should\s+be|shall\s+be|must\s+be|to\s+be\s+checked|"
     r"devrait\s+être|doit\s+être)\b",
     "normative_future", 1),

    # Rejected / refusal
    (r"\b(reject|refus|dismiss|decline|réfus)\b",
     "rejected", 3),

    # ── Additional patterns from negative signal detection ──
    # Fail / defect / bug / broken (severity: 3 — strong non-conformity signals)
    (r"\b(fail(ed|ure)?|defect|bug|broken)\b",
     "fail_defect", 3),

    # Missing / incomplete (severity: 2)
    (r"\b(missing|incomplete)\b",
     "missing_incomplete", 2),

    # Does not meet / comply / satisfy (severity: 3)
    (r"\b(does\s+not|do\s+not|doesn'?t|don'?t)\s*(meet|comply|satisf|fulfil)",
     "not_meet", 3),
    (r"\bnot\s+(met|satisfied|achieved|fulfilled|compliant)\b",
     "not_met", 3),

    # Non-conformity prefixes (non conform, non ok, non test, not ok)
    (r"\bnon\s*(conform|ok|test|verif|implement)\b",
     "non_conformity", 3),
    (r"\bnot?\s*(?:be\s+)?(ok|conform|test|implement|done|complete|met|satisf)\b",
     "negated_conformity", 3),

    # French: pas conforme / pas ok / pas terminé / pas fait / pas prêt / pas validé
    (r"\bpas\s+(conforme|ok|termin|fait|prêt|pret|valid)\b",
     "fr_pas_conforme", 3),
    (r"\bnon\s*conforme\b",
     "fr_non_conforme", 3),
    (r"\b(pas\s+(possible|capable|en\s+mesure))\b",
     "fr_cannot", 3),

    # Problème / souci (French problem/issue)
    (r"\b(problème|souci|préoccupation)\b",
     "fr_problem", 2),

    # Deviation / derogation / waiver (French)
    (r"\b(dérogation|écart|dispense)\b",
     "fr_deviation", 2),

    # Partial / limited (French)
    (r"\b(partiel|partielle|limité|limitée|solution\s+de\s+rechange)\b",
     "fr_partial", 2),

    # Conflict / contradiction (French)
    (r"\b(conflit|contradiction|incohérent|incohérence)\b",
     "fr_conflict", 2),

    # Refusé / refusée (French rejected)
    (r"\b(refus(é|ée|er))\b",
     "fr_rejected", 3),

    # Delay / late / retard
    (r"\b(delay|late|overdue|retard)\b",
     "delay", 1),

    # Under review / investigation
    (r"\b(under\s+(review|investigation)|being\s+(reviewed|investigated))\b",
     "under_review", 2),

    # TODO / TBD / TBA
    (r"\b(todo|tbd|tba)\b",
     "todo", 1),

    # No cybersecurity / no safety / no solution / no support / no capability
    (r"\bno\s+(cybersecurity|security|safety|solution|way|support|capability)\b",
     "no_noun", 2),

    # Still / remaining / outstanding (single words, broader)
    (r"\b(still|remaining|outstanding|not\s+yet)\b",
     "still_outstanding", 1),
    (r"\b(reste|encore|pas\s+encore)\b",
     "fr_remaining", 1),

    # Pending / waiting / blocked / not ready / not done / not available
    (r"\b(pending|wait|block|not\s+ready|not\s+done|not\s+available|"
     r"not\s+\w+(?:\s+\w+)?\s+yet)\b",
     "pending", 2),

    # KO (French rejection)
    (r"\bko\b", "ko", 3),

    # Error
    (r"\berror\b", "error", 2),
]


def _generate_ai_comment_ok(
    comment: str, matches: List[tuple], item_conformity: str
) -> str:
    """
    Generate an AI-style analysis comment explaining why an OK item's
    comment looks suspicious.
    """
    labels = sorted(set(l for l, _, _ in matches))
    score = sum(w for _, w, _ in matches)

    if "na_language" in labels:
        return (
            "⚠️ Le commentaire mentionne 'N/A' ou 'not applicable' mais le "
            "statut est OK. Incohérence logique — si l'exigence n'est pas "
            "applicable, le statut OK ne correspond pas au commentaire. "
            "Vérifier la cohérence entre le statut déclaré et le contenu."
        )
    if "cannot" in labels:
        return (
            "⚠️ Le commentaire indique une impossibilité ou incapacité technique "
            "('cannot', 'unable', 'impossible') alors que le statut FNR est OK. "
            "Incohérence logique — le commentaire décrit une non-conformité "
            "tandis que le statut déclare OK. Vérifier la cohérence."
        )
    if "not_responsible" in labels:
        return (
            "⚠️ Le fournisseur décline sa responsabilité dans le commentaire "
            "mais a marqué OK. Incohérence logique — si le fournisseur n'est "
            "pas responsable de cette exigence, le commentaire contredit le "
            "statut OK. Vérifier la cohérence."
        )
    if "rejected" in labels:
        return (
            "⚠️ Le commentaire contient un langage de refus/rejet alors que le "
            "statut est OK. Incohérence logique majeure — le commentaire "
            "décrit un refus mais le statut indique OK. Vérifier la cohérence."
        )
    if "pending_confirmation" in labels or "in_development" in labels:
        return (
            "⚠️ Le commentaire indique que ce point est encore en cours de "
            "développement ou en attente de confirmation, mais le statut est OK. "
            "Incohérence logique — le commentaire suggère que la conformité "
            "n'est pas encore validée. Vérifier la cohérence entre le statut "
            "déclaré et l'état réel."
        )
    if "needs_action" in labels or "stla_action" in labels:
        return (
            "⚠️ Le commentaire indique qu'une action est nécessaire (par le "
            "fournisseur ou STLA), mais le statut est déjà OK. Incohérence "
            "logique — si des actions sont encore requises, le commentaire "
            "contredit le statut OK. Vérifier la cohérence."
        )
    if "alternative_approach" in labels:
        return (
            "⚠️ Le commentaire mentionne une approche alternative, une déviation "
            "ou une substitution par rapport à l'exigence originale. Bien que le "
            "statut soit OK, l'approche alternative n'est pas documentée dans "
            "le statut. Vérifier la cohérence."
        )
    if "conflict" in labels:
        return (
            "⚠️ Le commentaire mentionne un conflit ou une contradiction avec "
            "une autre exigence. Le statut OK ne reflète pas cette situation "
            "décrite dans le commentaire — incohérence logique. "
            "Vérifier la cohérence."
        )
    if "partial_limited" in labels:
        return (
            "⚠️ Le commentaire suggère une conformité partielle ou limitée "
            "('partial', 'limited to', 'only for') alors que le statut est OK. "
            "Incohérence logique — le commentaire décrit des limitations "
            "qui contredisent un statut OK complet. Vérifier la cohérence."
        )
    if "temporary" in labels:
        return (
            "⚠️ Le commentaire mentionne une solution temporaire ou provisoire. "
            "Le statut OK ne mentionne pas cette condition temporaire — "
            "incohérence logique. Vérifier la cohérence entre le statut "
            "et le commentaire."
        )
    if "remaining" in labels:
        return (
            "⚠️ Le commentaire indique qu'il reste des éléments à compléter. "
            "Le statut OK ne reflète pas ces éléments restants — "
            "incohérence logique. Vérifier la cohérence."
        )

    # Generic fallback based on score
    if score >= 3:
        return (
            f"⚠️ Le commentaire contient plusieurs signaux ({', '.join(labels)}) "
            f"qui contredisent le statut OK déclaré par le fournisseur. "
            f"Incohérence logique — le commentaire ne correspond pas au "
            f"statut OK. Une revue humaine est recommandée pour vérifier "
            f"la cohérence."
        )
    if score >= 1:
        return (
            f"ℹ️ Le commentaire présente des signaux mineurs "
            f"({', '.join(labels)}) qui ne correspondent pas au statut OK. "
            f"Vérifier la cohérence entre le statut déclaré et le commentaire."
        )
    return ""


def _pattern_finding_for_item(item: ConformityItem) -> Optional[Dict]:
    """
    Pattern-based (regex) suspicion check for a single OK item.
    Returns a finding dict, or None if the comment raises no signal.
    """
    comment = item.comment.strip()
    conf_raw = item.conformity_raw.strip()

    cnorm = _normalize(comment)
    matches: List[tuple] = []
    for pattern, label, weight in _OK_SUSPICION_PATTERNS:
        m = re.search(pattern, cnorm)
        if m:
            matches.append((label, weight, m.group()))

    if not matches:
        return None

    score = sum(w for _, w, _ in matches)
    ai_comment = _generate_ai_comment_ok(comment, matches, conf_raw)

    return {
        "reqId": item.req_id,
        "reference": item.reference,
        "conformity": conf_raw,
        "comment": comment[:300],
        "score": score,
        "signals": sorted(set(l for l, _, _ in matches)),
        "matched": [t for _, _, t in matches],
        "aiComment": ai_comment,
        "severity": "error" if score >= 4 else "warning" if score >= 2 else "info",
        "source": "motifs",
    }


# ── LLM semantic deep analysis of OK responses ─────────────────────

_LLM_BATCH_SIZE = 25       # items per LLM call
_LLM_MAX_ITEMS = 150       # beyond this, remaining items fall back to patterns
_LLM_COMMENT_MAX_CHARS = 600

_LLM_SYSTEM_PROMPT = """Tu es un auditeur qualité senior spécialisé dans les matrices de conformité fournisseur (FNR) de l'industrie automobile.

Pour chaque exigence fournie, le fournisseur a déclaré le statut OK (conforme). Ta mission : juger si le COMMENTAIRE du fournisseur justifie réellement ce statut OK, ou s'il révèle en réalité un problème caché.

Rends un verdict pour CHAQUE exigence :
- "CONTRADICTION" : le commentaire décrit en réalité une non-conformité (refus, impossibilité, fonction absente ou non supportée, non applicable, hors périmètre, défaut connu...) → gravite "error"
- "PARTIEL" : conformité partielle, limitée, conditionnelle, avec déviation ou solution alternative non validée → gravite "warning"
- "EN_ATTENTE" : conformité non encore acquise (en cours, à confirmer, TBD, dépend d'une action, d'une livraison ou d'un essai futur...) → gravite "warning"
- "AMBIGU" : commentaire trop vague ou sans rapport pour justifier un OK → gravite "info"
- "COHERENT" : le commentaire confirme ou est compatible avec la conformité → gravite "none"

Règles :
- Les commentaires peuvent être en français ou en anglais.
- Un commentaire technique décrivant COMMENT l'exigence est satisfaite est COHERENT.
- Les commentaires du type « <domaine>: ok » (ex. « EE: ok », « SW: ok », « Touch: ok », « EE: ok SW: ok »), éventuellement accompagnés d'une date, sont des confirmations de conformité domaine par domaine : verdict COHERENT, jamais AMBIGU.
- De simples références (numéros de document, versions, dates, codes domaine) ne sont pas des problèmes.
- Ne signale AMBIGU que si le commentaire empêche réellement de comprendre pourquoi l'exigence serait conforme.
- "citation" : recopie exactement le fragment du commentaire (15 mots max) qui fonde ton verdict ; "" si COHERENT.
- "explication" : 1 à 2 phrases en français, précises et professionnelles.

Réponds UNIQUEMENT en JSON strict, sans texte autour :
{"resultats": [{"id": <int>, "verdict": "...", "gravite": "error|warning|info|none", "explication": "...", "citation": "..."}]}"""

_LLM_SEVERITY_SCORE = {"error": 5, "warning": 3, "info": 1}


def _llm_finding(item: ConformityItem, verdict: str, severity: str,
                 explication: str, citation: str) -> Dict:
    """Map one LLM verdict onto the standard finding schema."""
    icon = "ℹ️" if severity == "info" else "⚠️"
    ai_comment = f"{icon} {explication.strip()}"
    if citation:
        ai_comment += f" (extrait : « {citation.strip()} »)"
    return {
        "reqId": item.req_id,
        "reference": item.reference,
        "conformity": item.conformity_raw.strip(),
        "comment": item.comment.strip()[:300],
        "score": _LLM_SEVERITY_SCORE.get(severity, 1),
        "signals": ["analyse_ia", verdict.lower()],
        "matched": [citation] if citation else [],
        "aiComment": ai_comment,
        "severity": severity,
        "source": "ia",
    }


def _analyze_ok_deep_llm(items: List[ConformityItem]) -> Tuple[List[Dict], set]:
    """
    Semantic deep analysis of OK comments via the Azure OpenAI LLM.

    Sends the OK items (batched) to GPT and collects a verdict per item:
    CONTRADICTION / PARTIEL / EN_ATTENTE / AMBIGU / COHERENT.

    Returns (findings, analyzed_indices). Items whose batch failed are NOT
    in analyzed_indices — the caller falls back to pattern analysis for them.
    Returns ([], set()) when the LLM is not configured or unreachable.
    """
    import json as _json
    import logging as _logging

    if not items:
        return [], set()

    try:
        from app.config import (
            AZURE_OPENAI_API_KEY,
            AZURE_OPENAI_ENDPOINT,
            AZURE_OPENAI_LLM_DEPLOYMENT,
        )
        if not AZURE_OPENAI_API_KEY or not AZURE_OPENAI_ENDPOINT:
            return [], set()
        from app.embeddings import _get_client
        client = _get_client()
    except Exception as exc:
        _logging.warning(f"Deep-OK LLM unavailable (config/import): {exc}")
        return [], set()

    findings: List[Dict] = []
    analyzed: set = set()
    capped = items[:_LLM_MAX_ITEMS]

    for start in range(0, len(capped), _LLM_BATCH_SIZE):
        batch = capped[start:start + _LLM_BATCH_SIZE]
        lines = []
        for offset, item in enumerate(batch):
            idx = start + offset
            comment = item.comment.strip()[:_LLM_COMMENT_MAX_CHARS]
            conf = item.conformity_raw.strip() or "OK"
            lines.append(
                f"[{idx}] Exigence {item.req_id or '(sans id)'} — "
                f"statut déclaré : {conf}\nCommentaire : {comment}"
            )
        user_msg = (
            f"Analyse les {len(batch)} exigences suivantes "
            f"(toutes déclarées OK par le fournisseur) :\n\n"
            + "\n\n".join(lines)
        )

        try:
            response = client.chat.completions.create(
                model=AZURE_OPENAI_LLM_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": _LLM_SYSTEM_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
                temperature=0.0,
                max_tokens=4000,
                timeout=90,
            )
            text = (response.choices[0].message.content or "").strip()
            text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text)
            data = _json.loads(text)
            results = data.get("resultats", [])
        except Exception as exc:
            _logging.warning(
                f"Deep-OK LLM batch {start}-{start+len(batch)-1} failed: {exc}"
            )
            continue  # this batch falls back to patterns

        by_id = {r.get("id"): r for r in results if isinstance(r, dict)}
        for offset, item in enumerate(batch):
            idx = start + offset
            r = by_id.get(idx)
            if r is None:
                continue  # missing from response → pattern fallback
            analyzed.add(idx)
            verdict = str(r.get("verdict", "")).upper()
            severity = str(r.get("gravite", "none")).lower()
            if verdict == "COHERENT" or severity in ("none", ""):
                continue  # confirmed OK — no finding
            if severity not in ("error", "warning", "info"):
                severity = {
                    "CONTRADICTION": "error",
                    "PARTIEL": "warning",
                    "EN_ATTENTE": "warning",
                }.get(verdict, "info")
            if severity == "info":
                continue  # only real problems (error/warning) are reported
            findings.append(_llm_finding(
                item, verdict, severity,
                str(r.get("explication", "")).strip()
                or "Le commentaire ne justifie pas clairement le statut OK.",
                str(r.get("citation", "")).strip()[:120],
            ))

    return findings, analyzed


def analyze_ok_deep(analysis: ConformityAnalysis) -> List[Dict]:
    """
    Deep-analyze OK conformity items to detect hidden non-conformity signals.

    Two-stage analysis:
    1. Semantic LLM analysis (GPT via Azure OpenAI) — each OK comment is
       judged for real coherence with the declared OK status: contradiction,
       partial conformity, pending confirmation, ambiguity, or coherent.
    2. Pattern fallback — the proven regex suspicion library covers items
       the LLM could not analyze (not configured, unreachable, batch error,
       or beyond the per-run cap).

    Also flags OK items with no justifying comment (local check, no LLM).
    Sets analysis.ok_deep_method to "ia", "ia+motifs" or "motifs".
    """
    findings: List[Dict] = []

    # ── Local checks + collect items eligible for deep analysis ──
    deep_items: List[ConformityItem] = []
    for item in analysis.items:
        if item.conformity_category != "OK":
            continue

        comment = item.comment.strip()

        # OK without comment: nothing to analyze (info-level findings
        # are not reported — only real problems).
        if not comment:
            continue

        cnorm = _normalize(comment)

        # Skip pure domain code comments (domain assignments, not real comments)
        if re.match(
            r"^(sys|sw|ve|ee|me|od|opt|cg|tp|hw|mech|dq|fusa|ipm|all)"
            r"(\s*/\s*(sys|sw|ve|ee|me|od|opt|cg|tp|hw|mech|dq|fusa|ipm|all))*\s*$",
            cnorm,
        ):
            continue

        # Skip per-domain OK confirmations (e.g. "EE: ok", "Touch: ok",
        # "EE: ok SW: ok", "20260410 ME: ok") — these confirm conformity
        # domain by domain and are fully consistent with the OK status.
        if re.fullmatch(
            r"(?:[a-z0-9_.&/-]{1,15}\s*:\s*ok(?:ay)?|ok(?:ay)?|\d{2,8}|[\s,;/&+.-])+",
            cnorm,
        ):
            continue

        deep_items.append(item)

    # ── Stage 1: semantic LLM analysis ──
    llm_findings, analyzed_idx = _analyze_ok_deep_llm(deep_items)
    findings.extend(llm_findings)

    # ── Stage 2: pattern fallback for items the LLM did not cover ──
    remaining = [it for i, it in enumerate(deep_items) if i not in analyzed_idx]
    for item in remaining:
        f = _pattern_finding_for_item(item)
        if f:
            findings.append(f)

    if analyzed_idx and not remaining:
        analysis.ok_deep_method = "ia"
    elif analyzed_idx:
        analysis.ok_deep_method = "ia+motifs"
    else:
        analysis.ok_deep_method = "motifs"

    # Only real problems are reported — drop info-level findings
    # (whatever their source: LLM or pattern fallback).
    findings = [f for f in findings if f.get("severity") in ("error", "warning")]

    # Sort: most severe first, then score descending
    _sev_rank = {"error": 0, "warning": 1}
    findings.sort(key=lambda f: (_sev_rank.get(f.get("severity"), 2), -f.get("score", 0)))

    analysis.ok_deep_findings = findings
    # Backward compat: populate inconsistencies with the same unified findings
    analysis.inconsistencies = findings
    return findings


# ── Pie chart generation ────────────────────────────────────────────

# Color mapping for chart
_CHART_COLORS = {
    "OK": "#28a745",        # Green
    "NOK": "#dc3545",       # Red
    "NA": "#6c757d",        # Gray
    "EMPTY": "#e9ecef",     # Light gray
}


def _generate_svg_pie_chart(labels: list, sizes: list, colors: list,
                             title: str, total: int) -> str:
    """
    Generate a pie chart as SVG (pure Python, no matplotlib needed).
    Returns the SVG as a string.
    """
    import math

    cx, cy = 200, 200
    r = 130
    r_label = 165

    svg_parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="500" height="480" '
        f'viewBox="0 0 500 480">',
        f'<rect width="500" height="480" fill="white"/>',
        f'<text x="250" y="25" text-anchor="middle" font-size="14" '
        f'font-weight="bold" font-family="Arial" fill="#003366">'
        f'{_xml_escape(title)}</text>',
        f'<text x="250" y="45" text-anchor="middle" font-size="11" '
        f'font-family="Arial" fill="#666">({total} exigences)</text>',
    ]

    start_angle = -90.0  # Start at top (12 o'clock)
    for i, (label, size, color) in enumerate(zip(labels, sizes, colors)):
        if size == 0:
            continue
        pct = size / total * 100
        angle_span = (size / total) * 360.0
        end_angle = start_angle + angle_span

        # Calculate arc points
        x1 = cx + r * math.cos(math.radians(start_angle))
        y1 = cy + r * math.sin(math.radians(start_angle))
        x2 = cx + r * math.cos(math.radians(end_angle))
        y2 = cy + r * math.sin(math.radians(end_angle))

        large_arc = 1 if angle_span > 180 else 0

        # Pie slice path
        path = (
            f'M {cx},{cy} L {x1:.1f},{y1:.1f} '
            f'A {r},{r} 0 {large_arc} 1 {x2:.1f},{y2:.1f} Z'
        )
        svg_parts.append(
            f'<path d="{path}" fill="{color}" stroke="white" stroke-width="1.5"/>'
        )

        # Label position (midpoint of arc)
        mid_angle = start_angle + angle_span / 2
        lx = cx + r_label * math.cos(math.radians(mid_angle))
        ly = cy + r_label * math.sin(math.radians(mid_angle))

        # Percentage text inside the slice
        tx = cx + (r * 0.6) * math.cos(math.radians(mid_angle))
        ty = cy + (r * 0.6) * math.sin(math.radians(mid_angle))

        svg_parts.append(
            f'<text x="{lx:.1f}" y="{ly:.1f}" text-anchor="middle" '
            f'dominant-baseline="central" font-size="11" font-weight="bold" '
            f'font-family="Arial" fill="#333">{label}</text>'
        )
        svg_parts.append(
            f'<text x="{tx:.1f}" y="{ty:.1f}" text-anchor="middle" '
            f'dominant-baseline="central" font-size="10" font-weight="bold" '
            f'font-family="Arial" fill="white">{pct:.1f}%</text>'
        )

        start_angle = end_angle

    # Legend
    legend_y = 420
    legend_x = 50
    for i, (label, size, color) in enumerate(zip(labels, sizes, colors)):
        if size == 0:
            continue
        lx = legend_x + (i % 4) * 110
        ly = legend_y + (i // 4) * 20
        svg_parts.append(
            f'<rect x="{lx}" y="{ly - 8}" width="12" height="12" fill="{color}"/>'
        )
        svg_parts.append(
            f'<text x="{lx + 16}" y="{ly}" font-size="10" font-family="Arial" '
            f'fill="#333">{label}: {size}</text>'
        )

    svg_parts.append('</svg>')
    return "\n".join(svg_parts)


def _xml_escape(text: str) -> str:
    """Escape XML special characters."""
    return (text.replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def generate_pie_chart(analysis: ConformityAnalysis) -> str:
    """
    Generate a Camembert (pie) chart of conformity status distribution.
    Returns base64-encoded PNG string (or SVG if matplotlib not available).
    Gracefully returns empty string if no data.
    """
    # Filter out EMPTY for the chart (show meaningful statuses)
    chart_data = {k: v for k, v in analysis.stats.items()
                  if k != "EMPTY" and v > 0}

    if not chart_data:
        # If no meaningful data, show all
        chart_data = {k: v for k, v in analysis.stats.items() if v > 0}

    if not chart_data:
        return ""

    labels = list(chart_data.keys())
    sizes = list(chart_data.values())
    colors = [_CHART_COLORS.get(label, "#adb5bd") for label in labels]
    total = sum(sizes)

    title = (
        f"Repartition des statuts de conformite FNR"
    )

    # Try matplotlib first (generates PNG)
    try:
        import matplotlib
        matplotlib.use("Agg")  # Non-interactive backend
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(8, 6), dpi=100)

        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            colors=colors,
            autopct=lambda pct: f"{pct:.1f}%\n({int(round(pct/100*sum(sizes)))})",
            startangle=90,
            textprops={"fontsize": 11},
        )

        for autotext in autotexts:
            autotext.set_fontsize(9)
            autotext.set_fontweight("bold")

        ax.set_title(
            f"Répartition des statuts de conformité FNR\n"
            f"({total} exigences — {analysis.file_name or analysis.sheet_name})",
            fontsize=13,
            fontweight="bold",
            pad=20,
        )

        ax.axis("equal")

        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buf.seek(0)

        chart_b64 = base64.b64encode(buf.read()).decode("utf-8")
        analysis.chart_base64 = chart_b64
        return chart_b64

    except ImportError:
        # matplotlib not available — use pure-Python SVG fallback
        svg = _generate_svg_pie_chart(labels, sizes, colors, title, total)
        chart_b64 = base64.b64encode(svg.encode("utf-8")).decode("utf-8")
        analysis.chart_base64 = chart_b64
        return chart_b64


# ── Report text generation ──────────────────────────────────────────

def generate_report_text(analysis: ConformityAnalysis) -> str:
    """
    Generate a human-readable text report of the conformity analysis.
    """
    lines: List[str] = []

    lines.append("=" * 70)
    lines.append("LEON — Rapport d'Analyse de Matrice de Conformité FNR")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"Fichier analysé : {analysis.file_name}")
    lines.append(f"Feuille : {analysis.sheet_name}")
    lines.append(f"Ligne d'en-tête : {analysis.header_row + 1}")
    lines.append(f"Première ligne de données : {analysis.data_start_row + 1}")
    lines.append(f"Nombre total d'exigences : {analysis.total_rows}")
    lines.append("")

    # Statistics
    lines.append("─" * 50)
    lines.append("RÉPARTITION DES STATUTS DE CONFORMITÉ")
    lines.append("─" * 50)
    for cat, count in sorted(analysis.stats.items(), key=lambda x: -x[1]):
        pct = (count / analysis.total_rows * 100) if analysis.total_rows else 0
        lines.append(f"  {cat:12s} : {count:4d} ({pct:5.1f}%)")
    lines.append("")

    # OK items
    ok_items = [item for item in analysis.items if item.conformity_category == "OK"]
    if ok_items:
        lines.append("─" * 50)
        lines.append(f"EXIGENCES CONFORMES (OK) — {len(ok_items)}")
        lines.append("─" * 50)
        for item in ok_items:
            comment_str = f" | Commentaire: {item.comment}" if item.comment else ""
            version_str = f" | Version: {item.version}" if item.version else ""
            lines.append(f"  ✅ {item.req_id}: {item.conformity_raw}{version_str}{comment_str}")
        lines.append("")

    # NOK items
    nok_items = [item for item in analysis.items if item.conformity_category == "NOK"]
    if nok_items:
        lines.append("─" * 50)
        lines.append(f"EXIGENCES NON CONFORMES (NOK) — {len(nok_items)}")
        lines.append("─" * 50)
        for item in nok_items:
            comment_str = f" | Commentaire: {item.comment}" if item.comment else ""
            version_str = f" | Version: {item.version}" if item.version else ""
            lines.append(f"  ❌ {item.req_id}: {item.conformity_raw}{version_str}{comment_str}")
        lines.append("")

    # NA items
    na_items = [item for item in analysis.items if item.conformity_category == "NA"]
    if na_items:
        lines.append("─" * 50)
        lines.append(f"EXIGENCES NON APPLICABLES (NA) — {len(na_items)}")
        lines.append("─" * 50)
        for item in na_items:
            comment_str = f" | Commentaire: {item.comment}" if item.comment else ""
            version_str = f" | Version: {item.version}" if item.version else ""
            lines.append(f"  ⬜ {item.req_id}: {item.conformity_raw}{version_str}{comment_str}")
        lines.append("")

    # Items needing review
    review_items = [item for item in analysis.items if item.needs_review]
    if review_items:
        lines.append("─" * 50)
        lines.append(f"EXIGENCES À VÉRIFIER MANUELLEMENT — {len(review_items)}")
        lines.append("─" * 50)
        for item in review_items:
            conf_str = f"Statut: {item.conformity_category} ('{item.conformity_raw}')"
            comment_str = f" | Commentaire: {item.comment}" if item.comment else ""
            lines.append(f"  🔍 {item.req_id}: {conf_str}{comment_str}")
        lines.append("")

    # Deep analysis findings (unified: covers all OK suspicion signals)
    if analysis.inconsistencies:
        lines.append("─" * 50)
        lines.append(f"ANALYSE APPROFONDIE DES RÉPONSES OK — {len(analysis.inconsistencies)} point(s) d'attention")
        lines.append("─" * 50)
        for inc in analysis.inconsistencies:
            sev = inc.get("severity", "warning")
            icon = "🔴" if sev == "error" else "🟡" if sev == "warning" else "ℹ️"
            signals = ", ".join(inc.get("signals", []))
            lines.append(f"  {icon} [{sev.upper()}] {inc.get('reqId', inc.get('req_id', ''))} (score: {inc.get('score', 0)})")
            if signals:
                lines.append(f"     Signaux: {signals}")
            if inc.get('conformity'):
                lines.append(f"     Conformité: '{inc['conformity']}'")
            if inc.get('comment'):
                lines.append(f"     Commentaire: '{inc['comment'][:200]}'")
            if inc.get('aiComment'):
                lines.append(f"     Analyse: {inc['aiComment']}")
            lines.append("")
    else:
        lines.append("─" * 50)
        lines.append("✅ ANALYSE APPROFONDIE DES OK — Aucun point d'attention détecté")
        lines.append("─" * 50)
        lines.append("")

    lines.append("=" * 70)
    lines.append("Fin du rapport — LEON Conformity Matrix Analyzer")
    lines.append("=" * 70)

    report = "\n".join(lines)
    analysis.report_text = report
    return report


# ── Full analysis pipeline ──────────────────────────────────────────

def analyze_conformity_matrix(filepath: str, file_name: str = "") -> ConformityAnalysis:
    """
    Complete pipeline: extract → classify → deep analyze OK → chart → report.

    Args:
        filepath: Path to the ODS or XLSX file.
        file_name: Display name for the file.

    Returns:
        ConformityAnalysis with all data, stats, findings, chart, and report.
    """
    # 1. Extract data
    analysis = extract_conformity_data(filepath, file_name)

    # 2. Deep-analyze OK responses for hidden non-conformity (unified analysis)
    #    This replaces the old separate detect_inconsistencies() + analyze_ok_deep()
    #    which were analyzing the same OK items with overlapping patterns.
    analyze_ok_deep(analysis)

    # 3. Generate pie chart
    generate_pie_chart(analysis)

    # 4. Generate report text
    generate_report_text(analysis)

    return analysis


def analysis_to_dict(analysis: ConformityAnalysis) -> dict:
    """Convert ConformityAnalysis to a JSON-serializable dict."""
    return {
        "fileName": analysis.file_name,
        "sheetName": analysis.sheet_name,
        "headerRow": analysis.header_row,
        "dataStartRow": analysis.data_start_row,
        "totalRows": analysis.total_rows,
        "sheetTotalRows": analysis.sheet_total_rows,
        "debugRepeatedRows": getattr(analysis, "_debug_repeated_rows", 0),
        "debugTotalRowElements": getattr(analysis, "_debug_total_row_elements", 0),
        "debugTotalExpandedRows": getattr(analysis, "_debug_total_expanded_rows", 0),
        "stats": analysis.stats,
        "columnMapping": analysis.column_mapping,
        "items": [
            {
                "rowIndex": item.row_index,
                "reqId": item.req_id,
                "reference": item.reference,
                "description": item.description,
                "conformityRaw": item.conformity_raw,
                "conformityCategory": item.conformity_category,
                "comment": item.comment,
                "version": item.version,
                "versionApplicable": item.version,  # Alias for clarity
                "columnSet": item.column_set,
                "needsReview": item.needs_review,
                "classificationConfidence": item.classification_confidence,
            }
            for item in analysis.items
        ],
        "inconsistencies": analysis.inconsistencies,
        "okDeepFindings": analysis.ok_deep_findings,
        "okDeepMethod": analysis.ok_deep_method,
        "chartBase64": analysis.chart_base64,
        "reportText": analysis.report_text,
        "summary": {
            "total": analysis.total_rows,
            "ok": analysis.stats.get("OK", 0),
            "nok": analysis.stats.get("NOK", 0),
            "na": analysis.stats.get("NA", 0),
            "empty": analysis.stats.get("EMPTY", 0),
            "inconsistencies": len(analysis.inconsistencies),
            "okDeepFindings": len(analysis.ok_deep_findings),
            "needsReview": sum(1 for item in analysis.items if item.needs_review),
        },
    }


# ═══════════════════════════════════════════════════════════════════
# MULTI-MATRIX COMPARISON
# ═══════════════════════════════════════════════════════════════════

@dataclass
class MatrixComparison:
    """Result of comparing two or more conformity matrices."""
    matrices: List[Dict] = field(default_factory=list)  # per-matrix summaries
    # Per-requirement comparison: req_id → {matrix_name → category}
    requirement_comparison: Dict[str, Dict[str, str]] = field(default_factory=dict)
    # Requirements that changed status between matrices
    status_changes: List[Dict] = field(default_factory=list)
    # Requirements present in one matrix but not the other
    missing_in: Dict[str, List[str]] = field(default_factory=dict)
    # Comparison chart (base64 PNG)
    chart_base64: str = ""
    # Comparison report text
    report_text: str = ""
    # Summary
    total_compared: int = 0
    total_changes: int = 0
    total_missing: int = 0


def compare_matrices(filepaths: List[str], file_names: Optional[List[str]] = None) -> MatrixComparison:
    """
    Compare two or more conformity matrices side by side.

    For each requirement ID found in any matrix, shows its status in each matrix.
    Detects:
    - Status changes (e.g., NOK→OK between versions)
    - Requirements present in one matrix but missing in another
    - Overall trend (improvement or regression)

    Args:
        filepaths: List of ODS/XLSX file paths to compare.
        file_names: Optional display names (defaults to file basename).

    Returns:
        MatrixComparison with per-requirement comparison, changes, and chart.
    """
    if not filepaths:
        raise ValueError("At least one file path required for comparison")
    if file_names is None:
        file_names = [Path(f).name for f in filepaths]

    comparison = MatrixComparison()

    # Analyze each matrix
    analyses: List[ConformityAnalysis] = []
    for fp, fn in zip(filepaths, file_names):
        analysis = extract_conformity_data(fp, fn)
        analyze_ok_deep(analysis)
        analyses.append(analysis)

        comparison.matrices.append({
            "fileName": fn,
            "sheetName": analysis.sheet_name,
            "totalRows": analysis.total_rows,
            "stats": analysis.stats,
            "inconsistencies": len(analysis.inconsistencies),
            "summary": {
                "ok": analysis.stats.get("OK", 0),
                "nok": analysis.stats.get("NOK", 0),
                "na": analysis.stats.get("NA", 0),
                "empty": analysis.stats.get("EMPTY", 0),
            },
        })

    # Build per-requirement comparison
    # Use the first column set (column_set=0) for each matrix to avoid duplicates
    all_req_ids: set = set()
    per_matrix_items: Dict[str, Dict[str, str]] = {}  # matrix_name → {req_id → category}

    for analysis in analyses:
        matrix_name = analysis.file_name
        per_matrix_items[matrix_name] = {}
        for item in analysis.items:
            if item.column_set == 0:  # Use first column set only
                per_matrix_items[matrix_name][item.req_id] = item.conformity_category
                all_req_ids.add(item.req_id)

    # Build comparison dict
    for req_id in sorted(all_req_ids):
        row: Dict[str, str] = {}
        for matrix_name in file_names:
            row[matrix_name] = per_matrix_items.get(matrix_name, {}).get(req_id, "MISSING")
        comparison.requirement_comparison[req_id] = row

    comparison.total_compared = len(all_req_ids)

    # Detect status changes (only meaningful for 2 matrices)
    if len(analyses) == 2:
        m1_name, m2_name = file_names[0], file_names[1]
        m1_items = per_matrix_items.get(m1_name, {})
        m2_items = per_matrix_items.get(m2_name, {})

        for req_id in sorted(all_req_ids):
            cat1 = m1_items.get(req_id, "MISSING")
            cat2 = m2_items.get(req_id, "MISSING")

            if cat1 == "MISSING" and cat2 != "MISSING":
                comparison.missing_in.setdefault(m1_name, []).append(req_id)
                comparison.total_missing += 1
            elif cat2 == "MISSING" and cat1 != "MISSING":
                comparison.missing_in.setdefault(m2_name, []).append(req_id)
                comparison.total_missing += 1
            elif cat1 != cat2 and cat1 != "MISSING" and cat2 != "MISSING":
                comparison.status_changes.append({
                    "reqId": req_id,
                    "from": cat1,
                    "to": cat2,
                    "matrix1": m1_name,
                    "matrix2": m2_name,
                    "improvement": _is_improvement(cat1, cat2),
                })
                comparison.total_changes += 1

    # Generate comparison chart
    _generate_comparison_chart(comparison, file_names)

    # Generate comparison report text
    _generate_comparison_report(comparison, file_names)

    return comparison


def _is_improvement(from_cat: str, to_cat: str) -> bool:
    """Check if a status change is an improvement."""
    # OK is best, then NA, then EMPTY, then NOK
    ranking = {"OK": 3, "NA": 2, "EMPTY": 0, "NOK": -1}
    return ranking.get(to_cat, 0) > ranking.get(from_cat, 0)


def _generate_comparison_chart(comparison: MatrixComparison, file_names: List[str]) -> None:
    """Generate a grouped bar chart comparing conformity status across matrices."""
    categories = ["OK", "NOK", "NA", "EMPTY"]
    n_matrices = len(file_names)
    n_categories = len(categories)

    # Build data matrix: rows=categories, cols=matrices
    data = [[0] * n_matrices for _ in range(n_categories)]
    for mi, matrix_info in enumerate(comparison.matrices):
        stats = matrix_info.get("stats", {})
        for ci, cat in enumerate(categories):
            data[ci][mi] = stats.get(cat, 0)

    color_map = {
        "OK": "#28a745", "NOK": "#dc3545", "NA": "#6c757d",
        "EMPTY": "#e9ecef",
    }

    # Try matplotlib first (generates PNG)
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np

        data_np = np.array(data)

        fig, ax = plt.subplots(figsize=(10, 6), dpi=100)

        x = np.arange(n_categories)
        width = 0.8 / n_matrices

        for mi, matrix_info in enumerate(comparison.matrices):
            name = matrix_info.get("fileName", f"Matrix {mi+1}")
            short_name = name if len(name) <= 20 else name[:17] + "..."
            offset = (mi - n_matrices / 2 + 0.5) * width
            bars = ax.bar(x + offset, data_np[:, mi], width, label=short_name, alpha=0.85)

            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width() / 2., height,
                            f"{int(height)}", ha="center", va="bottom", fontsize=7)

        ax.set_xlabel("Conformity Status", fontsize=11, fontweight="bold")
        ax.set_ylabel("Number of Requirements", fontsize=11, fontweight="bold")
        ax.set_title("Conformity Matrix Comparison", fontsize=13, fontweight="bold", pad=15)
        ax.set_xticks(x)
        ax.set_xticklabels(categories, fontsize=10)
        ax.legend(fontsize=9, loc="upper right")
        ax.grid(axis="y", alpha=0.3)

        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buf.seek(0)

        comparison.chart_base64 = base64.b64encode(buf.read()).decode("utf-8")
        return

    except ImportError:
        pass  # Fall through to SVG fallback

    # SVG fallback (pure Python, no matplotlib)
    svg_parts = [
        '<svg xmlns="http://www.w3.org/2000/svg" width="600" height="400" viewBox="0 0 600 400">',
        '<rect width="600" height="400" fill="white"/>',
        '<text x="300" y="25" text-anchor="middle" font-size="14" font-weight="bold" '
        'font-family="Arial" fill="#003366">Conformity Matrix Comparison</text>',
    ]

    chart_left = 60
    chart_top = 50
    chart_w = 500
    chart_h = 280
    bar_area_w = chart_w / n_categories
    bar_w = bar_area_w * 0.8 / n_matrices

    # Find max value for scaling
    max_val = max(max(row) for row in data) if data else 1
    max_val = max(max_val, 1)

    # Y-axis grid lines
    for i in range(5):
        y = chart_top + chart_h * (1 - i / 4)
        val = int(max_val * i / 4)
        svg_parts.append(f'<line x1="{chart_left}" y1="{y:.0f}" x2="{chart_left + chart_w}" y2="{y:.0f}" stroke="#e0e0e0" stroke-width="1"/>')
        svg_parts.append(f'<text x="{chart_left - 5}" y="{y + 3:.0f}" text-anchor="end" font-size="9" font-family="Arial" fill="#666">{val}</text>')

    # Bars
    for ci, cat in enumerate(categories):
        cx = chart_left + ci * bar_area_w + bar_area_w / 2
        for mi in range(n_matrices):
            val = data[ci][mi]
            h = (val / max_val) * chart_h if max_val > 0 else 0
            bx = cx - (n_matrices * bar_w) / 2 + mi * bar_w
            by = chart_top + chart_h - h
            color = color_map.get(cat, "#adb5bd")
            svg_parts.append(f'<rect x="{bx:.1f}" y="{by:.1f}" width="{bar_w:.1f}" height="{h:.1f}" fill="{color}" stroke="white" stroke-width="0.5"/>')
            if val > 0:
                svg_parts.append(f'<text x="{bx + bar_w/2:.1f}" y="{by - 3:.0f}" text-anchor="middle" font-size="8" font-family="Arial" fill="#333">{val}</text>')
        # Category label
        svg_parts.append(f'<text x="{cx:.0f}" y="{chart_top + chart_h + 15}" text-anchor="middle" font-size="10" font-family="Arial" fill="#333">{cat}</text>')

    # Legend
    legend_y = 370
    for mi, matrix_info in enumerate(comparison.matrices):
        name = matrix_info.get("fileName", f"Matrix {mi+1}")
        short_name = name if len(name) <= 25 else name[:22] + "..."
        lx = 50 + mi * 200
        svg_parts.append(f'<rect x="{lx}" y="{legend_y - 8}" width="12" height="12" fill="#003366"/>')
        svg_parts.append(f'<text x="{lx + 16}" y="{legend_y}" font-size="9" font-family="Arial" fill="#333">{_xml_escape(short_name)}</text>')

    svg_parts.append('</svg>')
    svg = "\n".join(svg_parts)
    comparison.chart_base64 = base64.b64encode(svg.encode("utf-8")).decode("utf-8")


def _generate_comparison_report(comparison: MatrixComparison, file_names: List[str]) -> None:
    """Generate a text report for the multi-matrix comparison."""
    lines: List[str] = []

    lines.append("=" * 70)
    lines.append("LEON — Rapport de Comparaison Multi-Matrices de Conformité")
    lines.append("=" * 70)
    lines.append("")

    # Per-matrix summary
    lines.append("─" * 50)
    lines.append("RÉSUMÉ PAR MATRICE")
    lines.append("─" * 50)
    for matrix_info in comparison.matrices:
        lines.append(f"\n  📊 {matrix_info['fileName']}")
        lines.append(f"     Feuille: {matrix_info['sheetName']}")
        lines.append(f"     Total: {matrix_info['totalRows']} exigences")
        summary = matrix_info.get("summary", {})
        lines.append(f"     OK: {summary.get('ok', 0)} | NOK: {summary.get('nok', 0)} | "
                     f"NA: {summary.get('na', 0)} | EMPTY: {summary.get('empty', 0)}")
        lines.append(f"     Incohérences IA: {matrix_info.get('inconsistencies', 0)}")
    lines.append("")

    # Comparison summary
    lines.append("─" * 50)
    lines.append("COMPARAISON")
    lines.append("─" * 50)
    lines.append(f"  Exigences comparées: {comparison.total_compared}")
    lines.append(f"  Changements de statut: {comparison.total_changes}")
    lines.append(f"  Exigences manquantes: {comparison.total_missing}")
    lines.append("")

    # Status changes
    if comparison.status_changes:
        lines.append("─" * 50)
        lines.append(f"CHANGEMENTS DE STATUT — {len(comparison.status_changes)}")
        lines.append("─" * 50)
        improvements = [c for c in comparison.status_changes if c["improvement"]]
        regressions = [c for c in comparison.status_changes if not c["improvement"]]

        if improvements:
            lines.append(f"\n  ✅ AMÉLIORATIONS ({len(improvements)}):")
            for change in improvements[:20]:
                lines.append(f"    {change['reqId']}: {change['from']} → {change['to']}")

        if regressions:
            lines.append(f"\n  ❌ RÉGRESSIONS ({len(regressions)}):")
            for change in regressions[:20]:
                lines.append(f"    {change['reqId']}: {change['from']} → {change['to']}")
        lines.append("")

    # Missing requirements
    if comparison.missing_in:
        lines.append("─" * 50)
        lines.append("EXIGENCES MANQUANTES")
        lines.append("─" * 50)
        for matrix_name, req_ids in comparison.missing_in.items():
            lines.append(f"\n  Absentes dans '{matrix_name}': {len(req_ids)} exigences")
            for req_id in req_ids[:20]:
                lines.append(f"    {req_id}")
        lines.append("")

    lines.append("=" * 70)
    lines.append("Fin du rapport — LEON Multi-Matrix Comparison")
    lines.append("=" * 70)

    comparison.report_text = "\n".join(lines)


def comparison_to_dict(comparison: MatrixComparison) -> dict:
    """Convert MatrixComparison to a JSON-serializable dict."""
    return {
        "matrices": comparison.matrices,
        "requirementComparison": {
            req_id: statuses
            for req_id, statuses in comparison.requirement_comparison.items()
        },
        "statusChanges": comparison.status_changes,
        "missingIn": comparison.missing_in,
        "chartBase64": comparison.chart_base64,
        "reportText": comparison.report_text,
        "totalCompared": comparison.total_compared,
        "totalChanges": comparison.total_changes,
        "totalMissing": comparison.total_missing,
    }