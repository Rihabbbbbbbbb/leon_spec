"""
PDF report generator for conformity matrix analysis.

Generates a professional PDF report with:
- Summary statistics
- Embedded Camembert (pie) chart
- Lists of OK / NOK / NA items with exact comments
- AI-detected inconsistencies with explanations

Uses fpdf2 (pure Python, no C extensions) for Azure Function compatibility.
"""
from __future__ import annotations

import base64
import io
from typing import Dict

try:
    from fpdf import FPDF
    _HAS_FPDF = True
except ImportError:
    _HAS_FPDF = False


def _clean(text: str) -> str:
    """Sanitize text for latin-1 compatibility (fpdf2 Helvetica limitation)."""
    if not text:
        return ""
    replacements = {
        "\u2014": "-", "\u2013": "-", "\u2018": "'", "\u2019": "'",
        "\u201c": '"', "\u201d": '"', "\u2026": "...", "\u00a0": " ",
        "\u2192": "->", "\u2264": "<=", "\u2265": ">=", "\u00b0": " deg ",
        "\u00b1": "+/-", "\u00d7": "x", "\u00f7": "/", "\u2122": "(TM)",
        "\u00a9": "(c)", "\u00ae": "(R)", "\u2022": "-", "\u25cf": "o",
        "\u2610": "[ ]", "\u2713": "v", "\u2717": "x", "\u2705": "OK",
        "\u274c": "NOK", "\u23f3": "...", "\u2b1c": "[NA]",
        "\ud83d\udd34": "[!]", "\ud83d\udfe1": "[!]", "\ud83d\udfe2": "OK",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _status_color(category: str) -> tuple:
    """Return RGB color for a conformity category."""
    return {
        "OK": (40, 167, 69),
        "NOK": (220, 53, 69),
        "NA": (108, 117, 125),
        "EMPTY": (233, 236, 239),
    }.get(category, (108, 117, 125))


def _severity_color(severity: str) -> tuple:
    """Return RGB color for an inconsistency severity."""
    return {
        "error": (220, 53, 69),
        "warning": (255, 193, 7),
    }.get(severity, (108, 117, 125))


def generate_conformity_pdf(analysis_dict: dict) -> bytes:
    """
    Generate a professional PDF conformity analysis report.

    Args:
        analysis_dict: The analysis result dict from analysis_to_dict()

    Returns:
        PDF file as bytes
    """
    if not _HAS_FPDF:
        raise ImportError("fpdf2 is not installed. Install with: pip install fpdf2")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(10, 10, 10)
    _PAGE_WIDTH = 190  # A4 width 210mm - 2*10mm margins
    pdf.add_page()

    # ── Header ──────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 10, _clean("LEON - Rapport d'Analyse de Conformite FNR"),
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 5, _clean(f"Fichier: {analysis_dict.get('fileName', 'N/A')}"),
             new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, _clean(f"Feuille: {analysis_dict.get('sheetName', 'N/A')}"),
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ── Summary box ─────────────────────────────────────────
    summary = analysis_dict.get("summary", {})
    stats = analysis_dict.get("stats", {})

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 8, _clean("RESUME"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)

    total = summary.get("total", 0)
    ok = summary.get("ok", 0)
    nok = summary.get("nok", 0)
    na = summary.get("na", 0)
    empty = summary.get("empty", 0)
    inc_count = summary.get("inconsistencies", 0)

    # Summary table
    col_w = _PAGE_WIDTH / 4
    row_h = 7

    # Header row
    pdf.set_fill_color(0, 51, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w, row_h, "Statut", border=1, fill=True, align="C")
    pdf.cell(col_w, row_h, "Nombre", border=1, fill=True, align="C")
    pdf.cell(col_w, row_h, "Pourcentage", border=1, fill=True, align="C")
    pdf.cell(col_w, row_h, "Couleur", border=1, fill=True, align="C")
    pdf.ln(row_h)

    # Data rows
    pdf.set_font("Helvetica", "", 9)
    status_rows = [
        ("OK", ok, (40, 167, 69)),
        ("NOK", nok, (220, 53, 69)),
        ("NA", na, (108, 117, 125)),
        ("EMPTY", empty, (233, 236, 239)),
    ]

    for status, count, color in status_rows:
        pct = f"{count/total*100:.1f}%" if total else "0%"
        pdf.set_text_color(0, 0, 0)
        pdf.cell(col_w, row_h, status, border=1, align="L")
        pdf.cell(col_w, row_h, str(count), border=1, align="C")
        pdf.cell(col_w, row_h, pct, border=1, align="C")
        pdf.set_fill_color(*color)
        pdf.cell(col_w, row_h, "", border=1, fill=True, align="C")
        pdf.ln(row_h)

    # Total row
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w, row_h, "TOTAL", border=1, align="L")
    pdf.cell(col_w, row_h, str(total), border=1, align="C")
    pdf.cell(col_w, row_h, "100%", border=1, align="C")
    pdf.cell(col_w, row_h, "", border=1, align="C")
    pdf.ln(row_h)

    pdf.ln(4)

    # ── Camembert pie chart (integrated in summary) ────────
    chart_b64 = analysis_dict.get("chartBase64", "")
    if chart_b64:
        try:
            chart_bytes = base64.b64decode(chart_b64)
            chart_io = io.BytesIO(chart_bytes)
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(0, 51, 102)
            pdf.cell(0, 8, _clean("CAMEMBERT - Repartition des statuts de conformite"),
                     new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
            pdf.image(chart_io, x=30, w=130)
            pdf.ln(4)
        except Exception:
            pdf.set_font("Helvetica", "I", 9)
            pdf.set_text_color(150, 150, 150)
            pdf.cell(0, 5, _clean("(Graphique non disponible)"),
                     new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 5, _clean("(Graphique camembert non disponible)"),
                 new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # ── Inconsistencies summary ────────────────────────────
    pdf.set_font("Helvetica", "B", 10)
    if inc_count > 0:
        pdf.set_text_color(220, 53, 69)
        pdf.cell(0, 6, _clean(f"INCOHERENCES DETECTEES: {inc_count}"),
                 new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.set_text_color(40, 167, 69)
        pdf.cell(0, 6, _clean("AUCUNE INCOHERENCE DETECTEE"),
                 new_x="LMARGIN", new_y="NEXT")

    # ── OK items flagged by AI deep analysis ───────────────
    ok_findings = analysis_dict.get("okDeepFindings", [])
    if ok_findings:
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(255, 140, 0)
        pdf.cell(0, 6, _clean(f"ANALYSE APPROFONDIE DES OK — {len(ok_findings)} point(s) d'attention"),
                 new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)
        for finding in ok_findings[:30]:
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(255, 140, 0)
            pdf.cell(0, 5, _clean(f"  [!] {finding.get('reqId', '')}"),
                     new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(80, 80, 80)
            conf = finding.get("conformity", "")
            comment = finding.get("comment", "")
            analysis = finding.get("aiComment", "")
            if conf:
                pdf.cell(0, 4, _clean(f"     Conformite: {conf}"),
                         new_x="LMARGIN", new_y="NEXT")
            if comment:
                pdf.set_x(pdf.l_margin)
                try:
                    pdf.multi_cell(0, 4, _clean(f"     Commentaire: {comment}"),
                                 new_x="LMARGIN", new_y="NEXT")
                except Exception:
                    pdf.set_x(pdf.l_margin)
                    pdf.cell(0, 4, _clean(f"     Commentaire: {comment[:100]}..."),
                             new_x="LMARGIN", new_y="NEXT")
            if analysis:
                pdf.set_text_color(180, 70, 0)
                pdf.set_x(pdf.l_margin)
                try:
                    pdf.multi_cell(0, 4, _clean(f"     Analyse: {analysis}"),
                                 new_x="LMARGIN", new_y="NEXT")
                except Exception:
                    pdf.set_x(pdf.l_margin)
                    pdf.cell(0, 4, _clean(f"     Analyse: {analysis[:100]}..."),
                             new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
    else:
        pdf.ln(1)
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(40, 167, 69)
        pdf.cell(0, 6, _clean("Tous les statuts OK sont coherents avec leurs commentaires."),
                 new_x="LMARGIN", new_y="NEXT")

    pdf.ln(3)

    # ── Items by category ──────────────────────────────────
    items = analysis_dict.get("items", [])

    for category, label, icon in [
        ("OK", "EXIGENCES CONFORMES (OK)", "OK"),
        ("NOK", "EXIGENCES NON CONFORMES (NOK)", "NOK"),
        ("NA", "EXIGENCES NON APPLICABLES (NA)", "NA"),
    ]:
        cat_items = [item for item in items if item.get("conformityCategory") == category]
        if not cat_items:
            continue

        pdf.add_page()
        pdf.set_font("Helvetica", "B", 12)
        color = _status_color(category)
        pdf.set_text_color(*color)
        pdf.cell(0, 8, _clean(f"{label} - {len(cat_items)} exigences"),
                 new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(0, 0, 0)

        for item in cat_items:
            req_id = item.get("reqId", "")
            conf = item.get("conformityRaw", "")
            comment = item.get("comment", "")
            ref = item.get("reference", "")
            version = item.get("version", "")

            # Requirement ID + conformity
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*color)
            pdf.cell(0, 5, _clean(f"  [{icon}] {req_id}"), new_x="LMARGIN", new_y="NEXT")

            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(60, 60, 60)
            pdf.cell(0, 4, _clean(f"     Conformite: {conf}"), new_x="LMARGIN", new_y="NEXT")
            if ref:
                pdf.cell(0, 4, _clean(f"     Reference: {ref}"), new_x="LMARGIN", new_y="NEXT")
            if version:
                pdf.cell(0, 4, _clean(f"     Version applicable: {version}"), new_x="LMARGIN", new_y="NEXT")
            if comment:
                # Wrap long comments — reset X to left margin and handle errors
                pdf.set_text_color(0, 0, 0)
                comment_clean = _clean(comment)
                pdf.set_x(pdf.l_margin)
                try:
                    pdf.multi_cell(0, 4, f"     Commentaire: {comment_clean}",
                                 new_x="LMARGIN", new_y="NEXT")
                except Exception:
                    # Fallback: truncate to avoid rendering issues
                    pdf.set_x(pdf.l_margin)
                    pdf.cell(0, 4, _clean(f"     Commentaire: {comment_clean[:100]}..."),
                             new_x="LMARGIN", new_y="NEXT")
            else:
                pdf.set_text_color(150, 150, 150)
                pdf.cell(0, 4, _clean("     Commentaire: (aucun)"),
                         new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

    # ── Inconsistencies ─────────────────────────────────────
    inconsistencies = analysis_dict.get("inconsistencies", [])
    if inconsistencies:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(0, 51, 102)
        pdf.cell(0, 8, _clean(f"INCOHERENCES DETECTEES PAR L'IA - {len(inconsistencies)}"),
                 new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        for inc in inconsistencies:
            severity = inc.get("severity", "warning")
            color = _severity_color(severity)

            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(*color)
            sev_label = "ERREUR" if severity == "error" else "AVERTISSEMENT"
            score_str = f" (score: {inc.get('score', 0)})" if inc.get('score') else ""
            pdf.cell(0, 5, _clean(f"  [{sev_label}] {inc.get('reqId', '')} - {inc.get('type', '')}{score_str}"),
                     new_x="LMARGIN", new_y="NEXT")

            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(60, 60, 60)
            conf = inc.get("conformity", "")
            comment = inc.get("comment", "")
            signals = inc.get("signals", "")
            matched = inc.get("matched", [])
            if conf:
                pdf.cell(0, 4, _clean(f"     Conformite: {conf}"),
                         new_x="LMARGIN", new_y="NEXT")
            if signals:
                pdf.cell(0, 4, _clean(f"     Signaux detectes: {signals}"),
                         new_x="LMARGIN", new_y="NEXT")
            if matched:
                pdf.cell(0, 4, _clean(f"     Mots-cles: {', '.join(matched)}"),
                         new_x="LMARGIN", new_y="NEXT")
            if comment:
                pdf.set_x(pdf.l_margin)
                try:
                    pdf.multi_cell(0, 4, _clean(f"     Commentaire: {comment}"),
                                 new_x="LMARGIN", new_y="NEXT")
                except Exception:
                    pdf.set_x(pdf.l_margin)
                    pdf.cell(0, 4, _clean(f"     Commentaire: {comment[:100]}..."),
                             new_x="LMARGIN", new_y="NEXT")

            pdf.set_text_color(0, 0, 0)
            explanation = inc.get("explanation", "")
            if explanation:
                pdf.set_x(pdf.l_margin)
                try:
                    pdf.multi_cell(0, 4, _clean(f"     Analyse IA: {explanation}"),
                                 new_x="LMARGIN", new_y="NEXT")
                except Exception:
                    pdf.set_x(pdf.l_margin)
                    pdf.cell(0, 4, _clean(f"     Analyse IA: {explanation[:100]}..."),
                             new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

    # ── Footer ──────────────────────────────────────────────
    pdf.ln(5)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, _clean("Genere par LEON - Conformity Matrix Analyzer"),
             new_x="LMARGIN", new_y="NEXT")

    # Output to bytes
    out = pdf.output()
    return bytes(out)


# ═══════════════════════════════════════════════════════════════════
# EXCEL REPORT (color-coded rows)
# ═══════════════════════════════════════════════════════════════════

def _category_fill(category: str):
    """Return an openpyxl PatternFill for a conformity category."""
    from openpyxl.styles import PatternFill
    return {
        "OK": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),        # Green
        "NOK": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),       # Red
        "NA": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),         # Gray
        "EMPTY": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),     # White
    }.get(category, PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))


def _category_font(category: str):
    """Return an openpyxl Font for a conformity category."""
    from openpyxl.styles import Font
    return {
        "OK": Font(color="006100", bold=False),
        "NOK": Font(color="9C0006", bold=True),
        "NA": Font(color="3F3F3F", bold=False),
        "EMPTY": Font(color="808080", bold=False),
    }.get(category, Font(color="000000", bold=False))


def generate_conformity_excel(analysis_dict: dict) -> bytes:
    """
    Generate a color-coded Excel report from the conformity analysis.

    Sheets:
    1. "Summary" — statistics table + inconsistency count
    2. "All Items" — every requirement with color-coded rows by status
    3. "Inconsistencies" — AI-detected issues with severity colors

    Returns XLSX file as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(color="003366", bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # ── Sheet 1: Summary ────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary["A1"] = "LEON — Conformity Matrix Analysis"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:D1")

    ws_summary["A3"] = "File:"
    ws_summary["B3"] = analysis_dict.get("fileName", "N/A")
    ws_summary["A4"] = "Sheet:"
    ws_summary["B4"] = analysis_dict.get("sheetName", "N/A")
    ws_summary["A5"] = "Total Requirements:"
    ws_summary["B5"] = analysis_dict.get("totalRows", 0)
    ws_summary["A6"] = "Inconsistencies:"
    ws_summary["B6"] = len(analysis_dict.get("inconsistencies", []))

    # Statistics table
    ws_summary["A8"] = "Status"
    ws_summary["B8"] = "Count"
    ws_summary["C8"] = "Percentage"
    ws_summary["D8"] = "Color"
    for col in ["A8", "B8", "C8", "D8"]:
        ws_summary[col].fill = header_fill
        ws_summary[col].font = header_font
        ws_summary[col].border = thin_border

    stats = analysis_dict.get("stats", {})
    total = analysis_dict.get("totalRows", 0)
    row_idx = 9
    for category in ["OK", "NOK", "NA", "EMPTY"]:
        count = stats.get(category, 0)
        if count == 0 and category not in ("OK", "NOK", "NA"):
            continue
        pct = f"{count/total*100:.1f}%" if total else "0%"
        ws_summary[f"A{row_idx}"] = category
        ws_summary[f"B{row_idx}"] = count
        ws_summary[f"C{row_idx}"] = pct
        ws_summary[f"D{row_idx}"] = ""
        # Color the row
        fill = _category_fill(category)
        font = _category_font(category)
        for col in ["A", "B", "C", "D"]:
            cell = ws_summary[f"{col}{row_idx}"]
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
        row_idx += 1

    # Total row
    ws_summary[f"A{row_idx}"] = "TOTAL"
    ws_summary[f"B{row_idx}"] = total
    ws_summary[f"C{row_idx}"] = "100%"
    for col in ["A", "B", "C", "D"]:
        cell = ws_summary[f"{col}{row_idx}"]
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Column widths
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 15

    # ── Camembert pie chart in Summary ─────────────────────
    # Add chart data at position A15 for the pie chart source
    chart_data_start = row_idx + 2
    pie_labels_col = "A"
    pie_values_col = "B"
    ws_summary[f"A{chart_data_start}"] = "Category"
    ws_summary[f"B{chart_data_start}"] = "Count"
    ws_summary[f"A{chart_data_start}"].font = Font(bold=True, size=9)
    ws_summary[f"B{chart_data_start}"].font = Font(bold=True, size=9)
    pie_row = chart_data_start + 1
    pie_categories = ["OK", "NOK", "NA", "EMPTY"]
    for cat in pie_categories:
        cnt = stats.get(cat, 0)
        if cnt > 0:
            ws_summary[f"A{pie_row}"] = cat
            ws_summary[f"B{pie_row}"] = cnt
            fill = _category_fill(cat)
            ws_summary[f"A{pie_row}"].fill = fill
            ws_summary[f"B{pie_row}"].fill = fill
            ws_summary[f"A{pie_row}"].font = _category_font(cat)
            ws_summary[f"B{pie_row}"].font = _category_font(cat)
            pie_row += 1

    if pie_row > chart_data_start + 1:
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint
        pie_chart = PieChart()
        pie_chart.title = "Répartition des statuts de conformité"
        pie_chart.width = 18
        pie_chart.height = 12

        data_ref = Reference(ws_summary,
                             min_col=2, min_row=chart_data_start,
                             max_row=pie_row - 1)
        cats_ref = Reference(ws_summary,
                             min_col=1, min_row=chart_data_start + 1,
                             max_row=pie_row - 1)
        pie_chart.add_data(data_ref, titles_from_data=True)
        pie_chart.set_categories(cats_ref)

        # Color the slices
        chart_colors = ["28a745", "dc3545", "6c757d", "e9ecef"]
        for i, color in enumerate(chart_colors):
            if i < len(pie_chart.series[0].data_points):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = color
                pie_chart.series[0].data_points.append(pt)

        ws_summary.add_chart(pie_chart, "F2")
    else:
        ws_summary["A" + str(chart_data_start + 1)] = "Aucune donnée pour le camembert"
        ws_summary["A" + str(chart_data_start + 1)].font = Font(italic=True, color="808080")

    # ── Sheet 2: All Items ──────────────────────────────────
    ws_items = wb.create_sheet("All Items")

    headers = ["Row", "Req ID", "Reference", "Description", "Conformity (raw)",
               "Category", "Comment", "Version Applicable", "Column Set"]
    for ci, header in enumerate(headers, 1):
        cell = ws_items.cell(row=1, column=ci, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    items = analysis_dict.get("items", [])
    for ri, item in enumerate(items, 2):
        row_data = [
            item.get("rowIndex", 0),
            item.get("reqId", ""),
            item.get("reference", ""),
            item.get("description", ""),
            item.get("conformityRaw", ""),
            item.get("conformityCategory", ""),
            item.get("comment", ""),
            item.get("version", ""),
            item.get("columnSet", 0),
        ]
        category = item.get("conformityCategory", "EMPTY")
        fill = _category_fill(category)
        font = _category_font(category)

        for ci, val in enumerate(row_data, 1):
            cell = ws_items.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border
            if ci == 4:  # Description column
                cell.alignment = wrap_align

    # Column widths
    col_widths = [8, 20, 30, 50, 20, 15, 50, 20, 12]
    for ci, width in enumerate(col_widths, 1):
        ws_items.column_dimensions[get_column_letter(ci)].width = width

    # Freeze header row
    ws_items.freeze_panes = "A2"

    # Auto-filter
    ws_items.auto_filter.ref = f"A1:I{len(items) + 1}"

    # ── Sheet 3: Inconsistencies ────────────────────────────
    inconsistencies = analysis_dict.get("inconsistencies", [])
    if inconsistencies:
        ws_inc = wb.create_sheet("Inconsistencies")

        inc_headers = ["Severity", "Type", "Req ID", "Conformity", "Score", "Signals", "Matched Keywords", "Comment", "AI Explanation"]
        for ci, header in enumerate(inc_headers, 1):
            cell = ws_inc.cell(row=1, column=ci, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for ri, inc in enumerate(inconsistencies, 2):
            severity = inc.get("severity", "warning")
            sev_fill = PatternFill(
                start_color="FFC7CE" if severity == "error" else "FFEB9C",
                end_color="FFC7CE" if severity == "error" else "FFEB9C",
                fill_type="solid",
            )
            sev_font = Font(
                color="9C0006" if severity == "error" else "9C6500",
                bold=severity == "error",
            )
            row_data = [
                severity.upper(),
                inc.get("type", ""),
                inc.get("reqId", inc.get("req_id", "")),
                inc.get("conformity", ""),
                inc.get("score", ""),
                inc.get("signals", ""),
                ", ".join(inc.get("matched", [])) if inc.get("matched") else "",
                inc.get("comment", ""),
                inc.get("explanation", ""),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws_inc.cell(row=ri, column=ci, value=val)
                cell.fill = sev_fill
                cell.font = sev_font
                cell.border = thin_border
                if ci in (8, 9):
                    cell.alignment = wrap_align

        inc_widths = [12, 25, 20, 20, 8, 20, 25, 50, 60]
        for ci, width in enumerate(inc_widths, 1):
            ws_inc.column_dimensions[get_column_letter(ci)].width = width
        ws_inc.freeze_panes = "A2"
        ws_inc.auto_filter.ref = f"A1:I{len(inconsistencies) + 1}"

    # ── Sheet 4: OK Deep Analysis ──────────────────────────
    ok_findings = analysis_dict.get("okDeepFindings", [])
    if ok_findings:
        ws_ok = wb.create_sheet("OK Deep Analysis")

        ok_headers = ["Severity", "Req ID", "Reference", "Conformity",
                       "Score", "Signals", "Matched Keywords", "Comment",
                       "AI Analysis"]
        for ci, header in enumerate(ok_headers, 1):
            cell = ws_ok.cell(row=1, column=ci, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for ri, finding in enumerate(ok_findings, 2):
            sev = finding.get("severity", "warning")
            sev_fill = PatternFill(
                start_color="FFC7CE" if sev == "error" else
                "FFEB9C" if sev == "warning" else "C6EFCE",
                end_color="FFC7CE" if sev == "error" else
                "FFEB9C" if sev == "warning" else "C6EFCE",
                fill_type="solid",
            )
            sev_font = Font(
                color="9C0006" if sev == "error" else
                "9C6500" if sev == "warning" else "006100",
                bold=sev == "error",
            )
            row_data = [
                sev.upper(),
                finding.get("reqId", ""),
                finding.get("reference", ""),
                finding.get("conformity", ""),
                finding.get("score", 0),
                ", ".join(finding.get("signals", [])),
                ", ".join(finding.get("matched", [])),
                finding.get("comment", ""),
                finding.get("aiComment", ""),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws_ok.cell(row=ri, column=ci, value=val)
                cell.fill = sev_fill
                cell.font = sev_font
                cell.border = thin_border
                if ci in (8, 9):
                    cell.alignment = wrap_align

        ok_widths = [12, 20, 20, 20, 8, 20, 25, 50, 60]
        for ci, width in enumerate(ok_widths, 1):
            ws_ok.column_dimensions[get_column_letter(ci)].width = width
        ws_ok.freeze_panes = "A2"
        ws_ok.auto_filter.ref = f"A1:I{len(ok_findings) + 1}"

    # ── Save to bytes ───────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════
# POWER BI DASHBOARD JSON
# ═══════════════════════════════════════════════════════════════════

def generate_powerbi_dataset(analysis_dict: dict) -> dict:
    """
    Generate a Power BI-compatible dataset JSON for real-time conformity dashboards.

    Returns a dict with:
    - "datasets": list of tables (StatusSummary, Items, Inconsistencies)
    - "pushUrl": placeholder for Power BI streaming dataset URL
    - "dashboardConfig": suggested Power BI visual configuration

    This can be pushed to Power BI REST API or used as a streaming dataset.
    """
    stats = analysis_dict.get("stats", {})
    total = analysis_dict.get("totalRows", 0)
    items = analysis_dict.get("items", [])
    inconsistencies = analysis_dict.get("inconsistencies", [])

    # Table 1: Status Summary (for pie/donut chart)
    status_summary = []
    for category in ["OK", "NOK", "NA", "EMPTY"]:
        count = stats.get(category, 0)
        if count > 0:
            pct = round(count / total * 100, 2) if total else 0
            status_summary.append({
                "Status": category,
                "Count": count,
                "Percentage": pct,
                "Color": {
                    "OK": "#28a745", "NOK": "#dc3545", "NA": "#6c757d",
                    "EMPTY": "#e9ecef",
                }.get(category, "#adb5bd"),
            })

    # Table 2: Items (for detailed table visual)
    items_table = []
    for item in items:
        items_table.append({
            "ReqID": item.get("reqId", ""),
            "Reference": item.get("reference", ""),
            "Description": item.get("description", ""),
            "ConformityRaw": item.get("conformityRaw", ""),
            "Category": item.get("conformityCategory", ""),
            "Comment": item.get("comment", ""),
            "VersionApplicable": item.get("version", ""),
            "ColumnSet": item.get("columnSet", 0),
            "NeedsReview": item.get("needsReview", False),
            "ClassificationConfidence": item.get("classificationConfidence", "high"),
            "FileName": analysis_dict.get("fileName", ""),
            "SheetName": analysis_dict.get("sheetName", ""),
        })

    # Table 3: Inconsistencies (for alert visual)
    inc_table = []
    for inc in inconsistencies:
        inc_table.append({
            "Severity": inc.get("severity", "warning"),
            "Type": inc.get("type", ""),
            "ReqID": inc.get("reqId", inc.get("req_id", "")),
            "Conformity": inc.get("conformity", ""),
            "Comment": inc.get("comment", ""),
            "Explanation": inc.get("explanation", ""),
            "FileName": analysis_dict.get("fileName", ""),
        })

    # Power BI dataset definition (for REST API push)
    dataset_def = {
        "name": f"LEON_Conformity_{analysis_dict.get('fileName', 'matrix').replace('.', '_')}",
        "tables": [
            {
                "name": "StatusSummary",
                "columns": [
                    {"name": "Status", "dataType": "string"},
                    {"name": "Count", "dataType": "Int64"},
                    {"name": "Percentage", "dataType": "Double"},
                    {"name": "Color", "dataType": "string"},
                    {"name": "FileName", "dataType": "string"},
                    {"name": "Timestamp", "dataType": "DateTime"},
                ],
            },
            {
                "name": "Items",
                "columns": [
                    {"name": "ReqID", "dataType": "string"},
                    {"name": "Reference", "dataType": "string"},
                    {"name": "Description", "dataType": "string"},
                    {"name": "ConformityRaw", "dataType": "string"},
                    {"name": "Category", "dataType": "string"},
                    {"name": "Comment", "dataType": "string"},
                    {"name": "VersionApplicable", "dataType": "string"},
                    {"name": "ColumnSet", "dataType": "Int64"},
                    {"name": "FileName", "dataType": "string"},
                    {"name": "SheetName", "dataType": "string"},
                ],
            },
            {
                "name": "Inconsistencies",
                "columns": [
                    {"name": "Severity", "dataType": "string"},
                    {"name": "Type", "dataType": "string"},
                    {"name": "ReqID", "dataType": "string"},
                    {"name": "Conformity", "dataType": "string"},
                    {"name": "Comment", "dataType": "string"},
                    {"name": "Explanation", "dataType": "string"},
                    {"name": "FileName", "dataType": "string"},
                ],
            },
        ],
    }

    # Suggested dashboard visual configuration
    dashboard_config = {
        "visuals": [
            {
                "type": "pie",
                "title": "Conformity Status Distribution",
                "table": "StatusSummary",
                "category": "Status",
                "values": "Count",
                "colorBy": "Color",
            },
            {
                "type": "card",
                "title": "Total Requirements",
                "table": "StatusSummary",
                "aggregation": "sum",
                "field": "Count",
            },
            {
                "type": "card",
                "title": "NOK Count",
                "table": "StatusSummary",
                "filter": {"Status": "NOK"},
                "field": "Count",
            },
            {
                "type": "card",
                "title": "Inconsistencies",
                "table": "Inconsistencies",
                "aggregation": "count",
            },
            {
                "type": "table",
                "title": "All Requirements",
                "table": "Items",
                "columns": ["ReqID", "Reference", "Category", "ConformityRaw", "Comment", "VersionApplicable"],
                "sortBy": "Category",
                "colorBy": "Category",
            },
            {
                "type": "table",
                "title": "AI Inconsistencies",
                "table": "Inconsistencies",
                "columns": ["Severity", "ReqID", "Type", "Conformity", "Comment", "Explanation"],
                "filter": {"Severity": "error"},
                "colorBy": "Severity",
            },
            {
                "type": "bar",
                "title": "Requirements by Status",
                "table": "StatusSummary",
                "category": "Status",
                "values": "Count",
                "colorBy": "Color",
            },
        ],
    }

    return {
        "dataset": dataset_def,
        "data": {
            "StatusSummary": status_summary,
            "Items": items_table,
            "Inconsistencies": inc_table,
        },
        "dashboardConfig": dashboard_config,
        "summary": analysis_dict.get("summary", {}),
        "fileName": analysis_dict.get("fileName", ""),
        "sheetName": analysis_dict.get("sheetName", ""),
        "timestamp": __import__("datetime").datetime.now().isoformat(),
    }