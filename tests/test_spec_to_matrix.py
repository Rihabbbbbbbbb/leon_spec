"""
Tests for the Spec → Conformity Matrix generator (app/qa/spec_to_matrix.py).

Covers:
  1. Template integrity — single 'new version' sheet, no macros, formulas
     and dropdowns preserved
  2. Requirement extraction — block anchors, DOORS/internal ID coalescing,
     inline table rows, shall/must statements, history mentions, dedup
  3. Matrix generation — cells written at the right place, template intact
  4. Full pipeline on the real ASU spec (skipped if the file is absent)
"""
import io
import sys
import zipfile
from pathlib import Path

import pytest

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from app.qa.spec_to_matrix import (
    TEMPLATE_PATH,
    DATA_START_ROW,
    COL_DESCRIPTION,
    COL_REQ_ID,
    Requirement,
    extract_requirements,
    generate_conformity_matrix,
    spec_to_matrix,
)


ASU_PATH = Path(__file__).resolve().parent.parent / "data" / "uploads" / \
    "00692_25_01250_ASU_Technical_Specification_SPX _1_.docx"


# ── 1. Template integrity ─────────────────────────────────────────

class TestTemplate:

    def test_template_exists(self):
        assert TEMPLATE_PATH.exists(), f"Template missing: {TEMPLATE_PATH}"

    def test_template_has_no_macros(self):
        names = zipfile.ZipFile(TEMPLATE_PATH).namelist()
        assert not any("vba" in n.lower() for n in names)

    def test_template_single_new_version_sheet(self):
        from openpyxl import load_workbook
        wb = load_workbook(TEMPLATE_PATH)
        assert wb.sheetnames == ["new version"]

    def test_template_headers_and_formulas(self):
        from openpyxl import load_workbook
        ws = load_workbook(TEMPLATE_PATH)["new version"]
        assert "Libell" in (ws.cell(row=9, column=COL_DESCRIPTION).value or "")
        assert "exigence" in (ws.cell(row=9, column=COL_REQ_ID).value or "")
        assert (ws["F5"].value or "").startswith("=COUNTIF")

    def test_template_keeps_dropdowns(self):
        from openpyxl import load_workbook
        ws = load_workbook(TEMPLATE_PATH)["new version"]
        assert len(ws.data_validations.dataValidation) > 0


# ── 2. Requirement extraction ─────────────────────────────────────

class TestExtraction:

    def test_inline_table_row(self):
        text = 'REF-ASU-CD-LIN-0001(0) | The FNR must provide a justification folder for the line interface | [LIN1]'
        reqs = extract_requirements(text)
        assert len(reqs) == 1
        assert reqs[0].req_id == "REF-ASU-CD-LIN-0001"
        assert "justification folder" in reqs[0].text

    def test_block_anchor_with_internal_ref(self):
        text = (
            "REQ-0937326  C\n"
            "Requirement Number (v) | Description of the requirement | Input requirement (v)\n"
            "REF-ASU-CD-EXIFUNC-003\n"
            "Att_Sdf@ ASIL_A(A)\n"
            "PSA_Comments@{{ VF087_V2\n"
            "VF_2831}} | The function shall provide feedback within 200 ms. | [SSD_AUE]\n"
        )
        reqs = extract_requirements(text)
        by_id = {r.req_id: r for r in reqs}
        assert "REQ-0937326" in by_id
        req = by_id["REQ-0937326"]
        assert "REF-ASU-CD-EXIFUNC-003" in req.text
        assert "200 ms" in req.text

    def test_multiline_description_between_pipes(self):
        text = (
            "REQ-0937358  B\n"
            "Requirement Number (v) | Description of the requirement | Input requirement (v)\n"
            "REF-ASU-CD-EXIFUNC-004\n"
            "PSA_Comments@{{ X }} | During Idle State,\n"
            " IF\n"
            "Command is equal to Activation THEN the function shall switch state | [SSD]\n"
        )
        reqs = extract_requirements(text)
        req = {r.req_id: r for r in reqs}["REQ-0937358"]
        assert "During Idle State" in req.text
        assert "shall switch state" in req.text

    def test_history_mention_gets_no_description_but_real_definition_wins(self):
        text = (
            "New requirements:\n"
            "REQ-1111111\n"
            "Some unrelated line.\n"
            "REQ-1111111  C\n"
            "Requirement Number (v) | Description of the requirement | Input requirement (v)\n"
            "REF-X-CD-T-001\n"
            "meta@{{ x }} | The system shall do the real thing correctly. | [UP]\n"
        )
        reqs = extract_requirements(text)
        matches = [r for r in reqs if r.req_id == "REQ-1111111"]
        assert len(matches) == 1
        assert "real thing" in matches[0].text

    def test_broken_ref_spacing_repaired(self):
        text = "REF- ASU-CD-MAINT-0017(0) | The unit must survive 5 cycles of assembly. | [M1]"
        reqs = extract_requirements(text)
        assert reqs[0].req_id == "REF-ASU-CD-MAINT-0017"

    def test_shall_without_id_kept(self):
        text = "The ASU must be possible to disassemble and reassemble 20 times minimum without deterioration."
        reqs = extract_requirements(text)
        assert len(reqs) == 1
        assert reqs[0].req_id == ""
        assert "20 times" in reqs[0].text

    def test_template_example_filtered(self):
        text = "REF-PSP-FRONT-AIRBAG-001 | It is mandatory to write a Requirement no like: (free to modify the example) | x"
        reqs = extract_requirements(text)
        assert all("mandatory to write" not in r.text for r in reqs)

    def test_dedup_by_id(self):
        text = (
            "REF-A-CD-X-001 | The device shall blink twice per second. | [U1]\n"
            "REF-A-CD-X-001 | The device shall blink twice per second. | [U1]\n"
        )
        reqs = extract_requirements(text)
        assert len(reqs) == 1

    def test_short_prose_with_must_not_captured(self):
        text = "You must see this."  # < 30 chars, no ID
        reqs = extract_requirements(text)
        assert len(reqs) == 0

    def test_upstream_ref_not_captured_as_requirement(self):
        """'id | desc | upstream' rows: the trailing upstream id must NOT
        become a matrix row of its own."""
        text = "REF-ASU-CD-MAINT-0022(0) | FILL_FAULT_INFO_FRAME shall be activated in all functional states | REQ-0508543 A"
        reqs = extract_requirements(text)
        ids = {r.req_id for r in reqs}
        assert "REF-ASU-CD-MAINT-0022" in ids
        assert "REQ-0508543" not in ids

    def test_desc_then_id_layout_captured(self):
        """'description | id' rows (no leading id): the trailing id IS the
        requirement's own identifier."""
        text = "The failure of a single primary component must not generate the failure mode | GEN-ALM-CDC-SDF_041(0)"
        reqs = extract_requirements(text)
        assert len(reqs) == 1
        assert reqs[0].req_id == "GEN-ALM-CDC-SDF_041"
        assert "single primary component" in reqs[0].text

    def test_multiline_cell_merged_into_one_requirement_bruit_case(self):
        """A description cell continuing over several lines (numbered
        methods) must stay ONE requirement, not split into several rows."""
        text = (
            "Requirement Number (v) | Description of the requirement | Input requirement (v)\n"
            "REF-ASU-CD-BRUIT-0004(1) | Two methods are proposed to validate random noises:\n"
            "Method 1: During the test, emitted noise must be compliant with the Zwicker un-stationary loudness L10<4 sones\n"
            "Method 2: The measurement of random noise should be lower than the absence of random noise chart curve no. 1 + 3dB\n"
            "REF-ASU-CD-BRUIT-0005(0) | The listening for random noise should culminate in a rating. | [N42]\n"
        )
        reqs = extract_requirements(text)
        by_id = {r.req_id: r for r in reqs}
        assert "REF-ASU-CD-BRUIT-0004" in by_id
        assert "Method 1" in by_id["REF-ASU-CD-BRUIT-0004"].text
        assert "Method 2" in by_id["REF-ASU-CD-BRUIT-0004"].text
        # No separate id-less rows for the Method lines
        assert all("Method 1" not in r.text for r in reqs if not r.req_id)
        # Next requirement untouched
        assert "REF-ASU-CD-BRUIT-0005" in by_id

    def test_multiline_cell_merged_maint_case(self):
        """Second paragraph of the same cell (after a blank line, closing
        with '| [M20]') must merge into the requirement, not become its
        own row."""
        text = (
            "APP-ASU-CD-MAINT-0016(0) | The ASU must be possible to disassemble and reassemble 5 times minimum without any deterioration of characteristics defined\n"
            "\n"
            "The ASU must be possible to disassemble and reassemble 20 times minimum without any deterioration of characteristics defined in this document | [M20]\n"
        )
        reqs = extract_requirements(text)
        assert len(reqs) == 1
        req = reqs[0]
        assert req.req_id == "APP-ASU-CD-MAINT-0016"
        assert "5 times" in req.text
        assert "20 times" in req.text

    def test_inline_rows_not_swallowed_by_preceding_block(self):
        """An anchor block must stop at the first self-contained inline
        requirement row instead of swallowing it."""
        text = (
            "REF-A-CD-BLOCK-001\n"
            "some block metadata\n"
            "APP-A-CD-PERF-0001(0) | The unit must allow a sound level between 105 and 118 dB. | [M2]\n"
            "APP-A-CD-PERF-0002(0) | The unit must hold 350 seconds disconnected from supply. | [M2]\n"
        )
        reqs = extract_requirements(text)
        ids = {r.req_id: r for r in reqs}
        assert "APP-A-CD-PERF-0001" in ids
        assert "APP-A-CD-PERF-0002" in ids
        assert "sound level" in ids["APP-A-CD-PERF-0001"].text


# ── 3. Matrix generation ──────────────────────────────────────────

class TestGeneration:

    def test_fill_cells_and_preserve_template(self):
        from openpyxl import load_workbook
        reqs = [
            Requirement("REQ-0000001", "The system shall do A."),
            Requirement("REF-X-CD-Y-002", "The system shall do B."),
            Requirement("", "The system shall do C without an ID."),
        ]
        data = generate_conformity_matrix(reqs, "test")
        ws = load_workbook(io.BytesIO(data))["new version"]

        assert ws.cell(row=DATA_START_ROW, column=COL_REQ_ID).value == "REQ-0000001"
        assert ws.cell(row=DATA_START_ROW, column=COL_DESCRIPTION).value == "The system shall do A."
        assert ws.cell(row=DATA_START_ROW + 2, column=COL_REQ_ID).value in (None, "")
        assert "do C" in ws.cell(row=DATA_START_ROW + 2, column=COL_DESCRIPTION).value
        # Header + stats formula untouched
        assert "Libell" in (ws.cell(row=9, column=1).value or "")
        assert (ws["F5"].value or "").startswith("=COUNTIF")
        # Supplier columns stay empty
        for col in (4, 5, 6, 7, 8, 9):
            assert ws.cell(row=DATA_START_ROW, column=col).value in (None, "")

    def test_output_is_valid_macro_free_xlsx(self):
        reqs = [Requirement("REQ-1", "The system shall exist and be testable.")]
        data = generate_conformity_matrix(reqs, "t")
        names = zipfile.ZipFile(io.BytesIO(data)).namelist()
        assert not any("vba" in n.lower() for n in names)


# ── 4. Full pipeline on the real ASU spec ─────────────────────────

class TestRealSpec:

    @pytest.fixture(scope="class")
    def asu_result(self):
        if not ASU_PATH.exists():
            pytest.skip("ASU spec not found")
        from app.qa.retrieval import extract_text_from_file
        text = extract_text_from_file(ASU_PATH)
        return spec_to_matrix(text, ASU_PATH.name)

    def test_extracts_many_requirements(self, asu_result):
        assert asu_result["requirementsCount"] >= 200
        assert asu_result["withIdCount"] >= 200

    def test_most_ids_have_descriptions(self, asu_result):
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(asu_result["xlsxBytes"]))["new version"]
        with_id_and_desc = 0
        with_id = 0
        for r in range(DATA_START_ROW, ws.max_row + 1):
            rid = ws.cell(row=r, column=COL_REQ_ID).value
            if rid:
                with_id += 1
                if ws.cell(row=r, column=COL_DESCRIPTION).value:
                    with_id_and_desc += 1
        assert with_id > 0
        # At least 80% of identified requirements must carry a description
        assert with_id_and_desc / with_id >= 0.80

    def test_known_requirement_present_and_correct(self, asu_result):
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(asu_result["xlsxBytes"]))["new version"]
        for r in range(DATA_START_ROW, ws.max_row + 1):
            if ws.cell(row=r, column=COL_REQ_ID).value == "REQ-0937326":
                desc = ws.cell(row=r, column=COL_DESCRIPTION).value or ""
                assert "200 ms" in desc
                return
        pytest.fail("REQ-0937326 not found in generated matrix")
